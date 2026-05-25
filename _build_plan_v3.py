"""
Synapse Project Plan v3 — aggressive 2-month dev + 1-month UAT
Node.js backend (no .NET migration). 2 full-stack AI-engineer devs (Vibecoding).
Module-level ownership. Start 18 May 2026. SP->Postgres adapter already DONE.
Adds Policy-Based Access Control module.
"""
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = r"D:\Nalashaa_Work\AI_Work\synapse-fullstack\Synapse_Project_Plan_v3.xlsx"

# Brand
NAVY="1B2A5B"; LIGHT_NAVY="E8ECF5"; ACCENT="F26B21"; GREEN="2E7D32"; RED="C62828"
AMBER="F9A825"; GRAY="78909C"; DGRAY="455A64"; BLUE="1565C0"; WHITE="FFFFFF"; SUB="D7DCEA"
HF=Font(name="Calibri",size=11,bold=True,color=WHITE)
TF=Font(name="Calibri",size=20,bold=True,color=NAVY)
STF=Font(name="Calibri",size=12,bold=True,color=DGRAY)
BF=Font(name="Calibri",size=10); BB=Font(name="Calibri",size=10,bold=True)
SF=Font(name="Calibri",size=9,color=DGRAY)
NFILL=PatternFill("solid",fgColor=NAVY); LNFILL=PatternFill("solid",fgColor=LIGHT_NAVY)
SUBFILL=PatternFill("solid",fgColor=SUB)
THIN=Side(border_style="thin",color="B0BEC5"); BORDER=Border(THIN,THIN,THIN,THIN)
C=Alignment(horizontal="center",vertical="center",wrap_text=True)
L=Alignment(horizontal="left",vertical="center",wrap_text=True)
LT=Alignment(horizontal="left",vertical="top",wrap_text=True)

wb=Workbook()

# =================== TAB 1 — COVER ===================
ws=wb.active; ws.title="Cover"; ws.sheet_view.showGridLines=False
ws.column_dimensions["A"].width=3; ws.column_dimensions["B"].width=30; ws.column_dimensions["C"].width=92
ws.merge_cells("B2:C2"); ws["B2"]="Synapse Integration Platform"; ws["B2"].font=TF
ws.merge_cells("B3:C3"); ws["B3"]="Project Plan v3  ·  Aggressive 2-Month Build + 1-Month UAT"; ws["B3"].font=STF
meta=[("Generated",date.today().isoformat()),("Owner","Kiran Suvarna (Nalashaa)"),
("Delivery window","Dev: 18 May – 17 Jul 2026  ·  UAT: 20 Jul – 14 Aug 2026  ·  Go-live: ~15 Aug 2026"),
("Team","2 full-stack AI engineers (Vibecoding) — Dev A & Dev B. Module ownership, no FE/BE split."),
("Backend stack","Node.js + Express + TypeScript (kept). LLD's .NET is reference architecture only — no migration."),
("Code path",r"D:\Nalashaa_Work\AI_Work\synapse-fullstack"),
("Plan version","v3.0 (supersedes docs/Synapse_Project_Plan_v1.xlsx)")]
r=5
for k,v in meta:
    ws.cell(r,2,k).font=BB; ws.cell(r,3,v).font=BF; ws.cell(r,2).alignment=L; ws.cell(r,3).alignment=L; r+=1
r+=1
ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=3)
ws.cell(r,2,"What changed from v1").font=STF; ws.cell(r,2).fill=LNFILL; r+=1
changes=[
("Backend","Node.js retained. No .NET rewrite — the existing Jira + SharePoint engine stays and we build on it."),
("Pace","Compressed to 2-month development + 1-month UAT. Aggressive but feasible with Vibecoding velocity."),
("Team model","2 full-stack devs own whole modules end-to-end (build + wire + test). No separate FE/BE/QA dev lanes."),
("New module","Policy-Based Access Control added — company policy -> user subscription -> level of access. High priority."),
("Already done","SharePoint -> Postgres adapter complete and tested locally. Marked Done in the plan (0 remaining effort).")]
for k,v in changes:
    ws.cell(r,2,"▸ "+k).font=BB; ws.cell(r,3,v).font=BF; ws.cell(r,2).alignment=LT; ws.cell(r,3).alignment=LT
    ws.row_dimensions[r].height=30; r+=1
r+=1
ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=3)
ws.cell(r,2,"Scope — modules in this release").font=STF; ws.cell(r,2).fill=LNFILL; r+=1
scope=[
("Hub engine","Message envelope, bus, router, inbox/outbox/DLQ, retry + circuit breaker (Node/TS per LLD)."),
("Adapters","SharePoint->Postgres (DONE), SharePoint->MSSQL, Jira->SharePoint (harden)."),
("Connection Wizard","Full 6-step builder incl. mapping (Step 4) and test/deploy (Step 6)."),
("Mapping Canvas","Drag-map + AI auto-map (Claude API) with confidence scoring."),
("Policy-Based Access","Company policy -> user subscription -> enforced access level. JWT + middleware + admin UI."),
("Monitoring","Health Dashboard, Trading/Message Monitor (full trace), Alerts (in-app + email)."),
("Platform","Connector Studio, Master Entity Catalog (basic), Credential Vault, Scheduler, Audit log, Multi-tenancy.")]
for k,v in scope:
    ws.cell(r,2,k).font=BB; ws.cell(r,3,v).font=BF; ws.cell(r,2).alignment=LT; ws.cell(r,3).alignment=LT
    ws.row_dimensions[r].height=26; r+=1
r+=1
ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=3)
ws.cell(r,2,"Tab guide  ·  Status legend").font=STF; ws.cell(r,2).fill=LNFILL; r+=1
guide=[("Plan","Module-level tasks. Update Status, %, and Done Date. On-Time auto-calculates."),
("Milestones","6 checkpoints M1-M6 with target dates and acceptance criteria."),
("Sprint Plan","4 dev sprints + stabilization + UAT. Capacity vs loaded hours auto-calculated."),
("Risks & Decisions","Top risks and open decisions, kept short.")]
for k,v in guide:
    ws.cell(r,2,k).font=BB; ws.cell(r,3,v).font=BF; ws.cell(r,2).alignment=LT; ws.cell(r,3).alignment=LT; r+=1
r+=1
for label,col in [("Not Started",GRAY),("In Progress",BLUE),("Blocked",RED),("Done",GREEN)]:
    ws.cell(r,2,label).font=Font(name="Calibri",size=10,bold=True,color=WHITE)
    ws.cell(r,2).fill=PatternFill("solid",fgColor=col); ws.cell(r,2).alignment=C; ws.cell(r,2).border=BORDER; r+=1
ws.cell(r+1,2,"Effort hours are Vibecoding-adjusted (already discounted for AI-assisted velocity).").font=SF
ws.merge_cells(start_row=r+1,start_column=2,end_row=r+1,end_column=3)

# =================== TAB 2 — PLAN ===================
pl=wb.create_sheet("Plan"); pl.sheet_view.showGridLines=False
HEAD=["ID","Sprint","Module","Task","Owner","Effort (h)","Status","%","Start","Due","Done Date","On-Time?","Notes"]
for i,h in enumerate(HEAD,1):
    c=pl.cell(1,i,h); c.font=HF; c.fill=NFILL; c.alignment=C; c.border=BORDER
W=[8,11,24,52,9,9,12,7,12,12,12,12,46]
for i,w in enumerate(W,1): pl.column_dimensions[get_column_letter(i)].width=w
pl.row_dimensions[1].height=30; pl.freeze_panes="D2"

# sprint windows
SW={
"S1":(date(2026,5,18),date(2026,5,29)),
"S2":(date(2026,6,1), date(2026,6,12)),
"S3":(date(2026,6,15),date(2026,6,26)),
"S4":(date(2026,6,29),date(2026,7,10)),
"Stab":(date(2026,7,13),date(2026,7,17)),
"UAT":(date(2026,7,20),date(2026,8,14)),
}
# (sprint, module, task, owner, effort, status, pct, done_date_or_None, notes)
ROWS=[
# Sprint 1
("S1","Hub Engine","MessageEnvelope + SHA-256 checksum + topic model (Node/TS per LLD §6)","Dev A",24,"In Progress",0.25,None,"Team began 18 May. Immutable envelope, idempotency key."),
("S1","Hub Engine","IntegrationBus + Router + Subscription registry (in-memory queue + fan-out)","Dev A",24,"Not Started",0,None,"Backpressure via bounded queue."),
("S1","Hub Engine","Persistence: Inbox / Outbox / DeadLetter / Idempotency tables + repos (Postgres)","Dev A",20,"Not Started",0,None,"LLD §8 schema, Drizzle."),
("S1","Reliability","Retry+backoff + circuit breaker + DLQ auto/manual replay","Dev A",20,"Not Started",0,None,"Per Architecture §11."),
("S1","Adapter","SharePoint -> Postgres (delta poll, schema introspect, DDL diff/apply, upsert)","Dev B",0,"Done",1.0,date(2026,5,17),"COMPLETE — tested locally before kickoff."),
("S1","Adapter","SharePoint -> MSSQL (extend SP->PG: MERGE upsert + sys.columns introspect)","Dev B",24,"Not Started",0,None,"Reuse SP source + type mapper."),
("S1","Frontend","Wire Dashboard + Registry to real APIs (remove mock data)","Dev B",16,"Not Started",0,None,"First real-data screens."),
# Sprint 2
("S2","Wizard","Connection Wizard 6-step backend + Step 4 (mapping) + Step 6 (test/deploy gate)","Dev A",28,"Not Started",0,None,"Hard test-before-publish gate."),
("S2","Mapping","Mapping Canvas + AI Auto-Map (Claude API) + confidence badges","Dev B",30,"Not Started",0,None,"Human-confirm required."),
("S2","Policy Access","Policy-Based Access Control — data model + policy engine (policies, permissions, subscription)","Dev A",24,"Not Started",0,None,"NEW. Company policy -> access level."),
("S2","Frontend","Wire Monitor + Connected + Push to real APIs","Dev B",18,"Not Started",0,None,"Replace remaining mocks."),
("S2","Vault","Credential decrypt endpoint + DB-connection credential type + test-connection","Dev A",12,"Not Started",0,None,"AES-256-GCM in place."),
# Sprint 3
("S3","Policy Access","Policy enforcement middleware + JWT login + admin UI (manage policies & user subscriptions)","Dev A",32,"Not Started",0,None,"NEW. Subscription -> enforced level of access."),
("S3","Studio","Connector Studio — define / test / publish connector templates (backend + UI)","Dev B",28,"Not Started",0,None,"Designer workflow."),
("S3","Alerts","Alerts — failure detection + in-app feed + email dispatch + acknowledge","Dev A",20,"Not Started",0,None,"Wire dormant queue."),
("S3","Monitoring","Health Dashboard — real metrics, volume + error charts, drill-down","Dev B",20,"Not Started",0,None,"Replace hardcoded SVG."),
("S3","Audit","Audit log — middleware + viewer (ties to Policy module)","Dev A",14,"Not Started",0,None,"Append-only, before/after diff."),
# Sprint 4
("S4","Monitoring","Trading / Message Monitor — full per-message trace + error drill-down + create-task","Dev B",24,"Not Started",0,None,"Expandable rows, mapping trace."),
("S4","Catalog","Master Entity Catalog (basic) — entity registry + field-level usage","Dev A",20,"Not Started",0,None,"Cross-adapter reference."),
("S4","Platform","Multi-tenancy / org enforcement (row-level org_id, integrated with Policy)","Dev A",16,"Not Started",0,None,"Data isolation."),
("S4","Scheduler","Scheduler wire + cron builder UI (BullMQ repeatable jobs)","Dev B",16,"Not Started",0,None,"Next-run preview."),
("S4","Adapter","Jira -> SharePoint parity verify + harden","Dev B",12,"Not Started",0,None,"Existing engine regression."),
("S4","Integration","End-to-end integration testing + bug bash (all modules)","Both",24,"Not Started",0,None,"Feature-complete gate."),
# Stabilization
("Stab","Hardening","Security review (vault, policy enforcement, OWASP) + perf test (parallel workers)","Both",20,"Not Started",0,None,"Pre-UAT gate."),
("Stab","Hardening","Regression suite + smoke + fixes","Both",16,"Not Started",0,None,"Stabilize."),
("Stab","Deploy","Deploy to UAT environment (Docker) + smoke + runbook","Dev A",12,"Not Started",0,None,"Code freeze."),
# UAT
("UAT","UAT","UAT planning + scenarios + test-data setup","BA/QA",24,"Not Started",0,None,"Scripts per module."),
("UAT","UAT","UAT cycle 1 execution","QA + Stakeholders",40,"Not Started",0,None,"Log defects."),
("UAT","UAT","Defect triage + fixes (cycle 1)","Both",40,"Not Started",0,None,"Devs on standby."),
("UAT","UAT","UAT cycle 2 + regression","QA",24,"Not Started",0,None,"Verify fixes."),
("UAT","UAT","UAT signoff + go-live readiness","BA + Stakeholders",8,"Not Started",0,None,"Final acceptance."),
]
RS=2; row=RS
spkeys=[]
first={}; last={}
for i,(sp,mod,task,owner,eff,status,pct,done,notes) in enumerate(ROWS,1):
    s,e=SW[sp]
    if sp not in first: first[sp]=row; spkeys.append(sp)
    last[sp]=row
    pl.cell(row,1,f"T-{i:02d}").font=BF
    pl.cell(row,2,sp).font=BF
    pl.cell(row,3,mod).font=BF
    pl.cell(row,4,task).font=BF
    pl.cell(row,5,owner).font=BF
    pl.cell(row,6,eff).font=BF
    pl.cell(row,7,status).font=BF
    pl.cell(row,8,pct).font=BF; pl.cell(row,8).number_format="0%"
    # SP->PG done before kickoff -> start/due reflect actual
    if done==date(2026,5,17):
        pl.cell(row,9,date(2026,5,12))
    else:
        pl.cell(row,9,s)
    pl.cell(row,10,e)
    pl.cell(row,9).number_format="yyyy-mm-dd"; pl.cell(row,10).number_format="yyyy-mm-dd"
    if done: pl.cell(row,11,done)
    pl.cell(row,11).number_format="yyyy-mm-dd"
    pl.cell(row,12,f'=IF(K{row}="","",IF(K{row}<=J{row},"On-Time","Late"))').font=BF
    pl.cell(row,13,notes).font=BF
    for col in range(1,14):
        cc=pl.cell(row,col); cc.border=BORDER
        cc.alignment=LT if col in (3,4,13) else C
    pl.row_dimensions[row].height=30
    row+=1
last_row=row-1

# subtotal per sprint
for sp in spkeys:
    pl.cell(row,3,f"{sp} subtotal").font=BB
    pl.cell(row,6,f"=SUM(F{first[sp]}:F{last[sp]})").font=BB; pl.cell(row,6).number_format="#,##0"
    for col in range(1,14): pl.cell(row,col).fill=SUBFILL; pl.cell(row,col).border=BORDER
    row+=1
pl.cell(row,3,"GRAND TOTAL (hrs)").font=HF
pl.cell(row,6,f"=SUM(F{RS}:F{last_row})").font=HF; pl.cell(row,6).number_format="#,##0"
for col in range(1,14): pl.cell(row,col).fill=NFILL; pl.cell(row,col).border=BORDER; pl.cell(row,col).font=HF

# validations + CF
dvs=DataValidation(type="list",formula1='"Not Started,In Progress,Blocked,Done"',allow_blank=True); dvs.add(f"G{RS}:G{last_row}"); pl.add_data_validation(dvs)
dvp=DataValidation(type="list",formula1='"0,0.25,0.5,0.75,1"',allow_blank=True); dvp.add(f"H{RS}:H{last_row}"); pl.add_data_validation(dvp)
for st,col in {"Not Started":GRAY,"In Progress":BLUE,"Blocked":RED,"Done":GREEN}.items():
    pl.conditional_formatting.add(f"G{RS}:G{last_row}",CellIsRule(operator="equal",formula=[f'"{st}"'],fill=PatternFill("solid",fgColor=col),font=Font(color=WHITE,bold=True,size=10)))
pl.conditional_formatting.add(f"L{RS}:L{last_row}",CellIsRule(operator="equal",formula=['"On-Time"'],fill=PatternFill("solid",fgColor=GREEN),font=Font(color=WHITE,bold=True,size=10)))
pl.conditional_formatting.add(f"L{RS}:L{last_row}",CellIsRule(operator="equal",formula=['"Late"'],fill=PatternFill("solid",fgColor=RED),font=Font(color=WHITE,bold=True,size=10)))
pl.conditional_formatting.add(f"H{RS}:H{last_row}",DataBarRule(start_type="num",start_value=0,end_type="num",end_value=1,color="4472C4",showValue=True))

# =================== TAB 3 — MILESTONES ===================
ms=wb.create_sheet("Milestones"); ms.sheet_view.showGridLines=False
MH=["ID","Milestone","Target Date","Status","Owner","Acceptance Criteria","Actual Date","Slip (days)"]
for i,h in enumerate(MH,1):
    c=ms.cell(1,i,h); c.font=HF; c.fill=NFILL; c.alignment=C; c.border=BORDER
for i,w in enumerate([8,46,14,14,16,64,14,11],1): ms.column_dimensions[get_column_letter(i)].width=w
ms.row_dimensions[1].height=30; ms.freeze_panes="A2"
MIL=[
("M1","Hub engine live + adapters working + Dashboard/Registry on real data",date(2026,5,29),"Dev A",
 "Message flows Inbox->Bus->Router->Outbox; DLQ on forced failure; SP->PG + SP->MSSQL write live; Dashboard & Registry show real runs."),
("M2","Connection Wizard end-to-end + Mapping Canvas with AI auto-map",date(2026,6,12),"Dev A",
 "Operator builds an integration through all 6 steps; AI auto-map suggests with confidence; test-before-publish gate enforced."),
("M3","Policy-Based Access Control GA + Alerts + Health Dashboard",date(2026,6,26),"Dev A",
 "Company policy -> user subscription -> enforced access level verified across roles; failure alert reaches in-app + email; Health Dashboard live on real metrics."),
("M4","Feature-complete — all modules integrated, E2E passing",date(2026,7,10),"Both",
 "All Plan modules Done; end-to-end test suite green; no open Sev-1/2 defects."),
("M5","Code freeze — security/perf cleared, deployed to UAT",date(2026,7,17),"Both",
 "Security review passed (vault + policy enforcement); perf test meets target; build deployed to UAT env with smoke pass + runbook."),
("M6","UAT signoff & go-live readiness",date(2026,8,14),"BA + Stakeholders",
 "UAT cycles 1-2 complete; all signoff scenarios passed; go-live checklist approved."),
]
for i,(mid,m,tgt,owner,ac) in enumerate(MIL,2):
    ms.cell(i,1,mid).font=BB; ms.cell(i,2,m).font=BF; ms.cell(i,3,tgt).number_format="yyyy-mm-dd"
    ms.cell(i,4,"Not Started").font=BF; ms.cell(i,5,owner).font=BF; ms.cell(i,6,ac).font=BF
    ms.cell(i,7,None).number_format="yyyy-mm-dd"
    ms.cell(i,8,f'=IF(G{i}="","",NETWORKDAYS(C{i},G{i})-1)').font=BF
    for col in range(1,9):
        ms.cell(i,col).border=BORDER; ms.cell(i,col).alignment=LT if col in (2,6) else C
    ms.row_dimensions[i].height=54
dvm=DataValidation(type="list",formula1='"Not Started,In Progress,At Risk,Achieved,Missed"',allow_blank=True); dvm.add(f"D2:D{len(MIL)+1}"); ms.add_data_validation(dvm)
for st,col in {"Not Started":GRAY,"In Progress":BLUE,"At Risk":AMBER,"Achieved":GREEN,"Missed":RED}.items():
    ms.conditional_formatting.add(f"D2:D{len(MIL)+1}",CellIsRule(operator="equal",formula=[f'"{st}"'],fill=PatternFill("solid",fgColor=col),font=Font(color=WHITE,bold=True,size=10)))

# =================== TAB 4 — SPRINT PLAN ===================
sp=wb.create_sheet("Sprint Plan"); sp.sheet_view.showGridLines=False
SH=["Sprint","Start","End","Goal","Capacity (h)","Loaded (h)","Utilization","Milestone"]
for i,h in enumerate(SH,1):
    c=sp.cell(1,i,h); c.font=HF; c.fill=NFILL; c.alignment=C; c.border=BORDER
for i,w in enumerate([10,12,12,58,13,12,12,12],1): sp.column_dimensions[get_column_letter(i)].width=w
sp.row_dimensions[1].height=30; sp.freeze_panes="A2"
SPR=[
("S1",SW["S1"],"Hub engine + reliability · SP->MSSQL adapter · Dashboard/Registry wired",144,"M1"),
("S2",SW["S2"],"Connection Wizard · Mapping Canvas + AI auto-map · Policy model · Vault",144,"M2"),
("S3",SW["S3"],"Policy enforcement + admin UI · Connector Studio · Alerts · Health Dashboard · Audit",144,"M3"),
("S4",SW["S4"],"Trading Monitor · Entity Catalog · Multi-tenancy · Scheduler · E2E bug bash",144,"M4"),
("Stab",SW["Stab"],"Security + perf review · regression · deploy to UAT (code freeze)",72,"M5"),
("UAT",SW["UAT"],"UAT cycles 1-2 · defect fixing · signoff · go-live readiness",240,"M6"),
]
for i,(spk,(s,e),goal,cap,mil) in enumerate(SPR,2):
    sp.cell(i,1,spk).font=BB; sp.cell(i,2,s).number_format="yyyy-mm-dd"; sp.cell(i,3,e).number_format="yyyy-mm-dd"
    sp.cell(i,4,goal).font=BF; sp.cell(i,5,cap).font=BB
    sp.cell(i,6,f'=SUMIFS(Plan!F{RS}:F{last_row},Plan!B{RS}:B{last_row},A{i})').number_format="#,##0"
    sp.cell(i,7,f"=IFERROR(F{i}/E{i},0)").number_format="0%"
    sp.cell(i,8,mil).font=BF
    for col in range(1,9): sp.cell(i,col).border=BORDER; sp.cell(i,col).alignment=LT if col==4 else C
    sp.row_dimensions[i].height=42
sp.conditional_formatting.add(f"G2:G{len(SPR)+1}",CellIsRule(operator="greaterThan",formula=["1"],fill=PatternFill("solid",fgColor=RED),font=Font(color=WHITE,bold=True,size=10)))
sp.conditional_formatting.add(f"G2:G{len(SPR)+1}",CellIsRule(operator="between",formula=["0.9","1"],fill=PatternFill("solid",fgColor=AMBER),font=Font(color=WHITE,bold=True,size=10)))
sp.conditional_formatting.add(f"G2:G{len(SPR)+1}",CellIsRule(operator="lessThan",formula=["0.9"],fill=PatternFill("solid",fgColor=GREEN),font=Font(color=WHITE,bold=True,size=10)))
sp.cell(len(SPR)+3,1,"Capacity: 2 devs × ~36 productive h/week (Vibecoding) = 144 h per 2-week sprint. Stab = 1 week. UAT capacity blends devs + QA + stakeholders.").font=SF
sp.merge_cells(start_row=len(SPR)+3,start_column=1,end_row=len(SPR)+3,end_column=8)

# =================== TAB 5 — RISKS & DECISIONS ===================
rk=wb.create_sheet("Risks & Decisions"); rk.sheet_view.showGridLines=False
RH=["ID","Type","Item","Likelihood","Impact","Mitigation / Action","Owner","Due","Status"]
for i,h in enumerate(RH,1):
    c=rk.cell(1,i,h); c.font=HF; c.fill=NFILL; c.alignment=C; c.border=BORDER
for i,w in enumerate([7,10,52,12,10,56,12,13,12],1): rk.column_dimensions[get_column_letter(i)].width=w
rk.row_dimensions[1].height=30; rk.freeze_panes="A2"
RISK=[
("R1","Risk","Two-month window is aggressive; any module slip cascades","H","H","Weekly burn-down vs Sprint capacity; cut Catalog/Studio to 'basic' first if behind; protect Policy module + adapters.","Kiran",date(2026,5,29),"Open"),
("R2","Risk","Policy-Based Access is new and security-sensitive","M","H","Spike the policy model in S2 before enforcement in S3; security review at Stab; least-privilege defaults.","Dev A",date(2026,6,12),"Open"),
("R3","Risk","Vibecoding velocity assumption (144h/sprint) may not hold for hub internals","M","M","Track actual vs planned each sprint; rebaseline after S1; keep S4 light as buffer.","Kiran",date(2026,5,29),"Open"),
("R4","Risk","Frontend mock->real swap surfaces API contract gaps","M","M","Generate TS types from backend DTOs; wire one screen end-to-end in S1 as the pattern.","Dev B",date(2026,5,29),"Open"),
("R5","Risk","Playwright/MFA adapters fragile to client config","M","M","Smoke each adapter after deploy; session persistence + re-login fallback.","Dev B",date(2026,7,17),"Open"),
("R6","Risk","UAT defects exceed 1-month fix capacity","M","H","Bug-bash in S4 before freeze; both devs reserved for UAT defect fixing; triage Sev-1/2 only for go-live.","Both",date(2026,8,7),"Open"),
("D1","Decision","AI provider for auto-mapping (Claude / OpenAI / local)","—","M","Default Claude API; confirm by S2 start on cost × accuracy.","Dev B",date(2026,6,1),"Open"),
("D2","Decision","Policy model — RBAC roles vs full ABAC (attribute) policies","—","H","Recommend policy = named permission-set users subscribe to (RBAC+scope). Confirm S2.","Kiran",date(2026,6,1),"Open"),
("D3","Decision","UAT environment host — local box / VM / cloud","—","M","Pick by S3 so Stab deploy is ready. Default: in-office VM + Docker.","Dev A",date(2026,6,15),"Open"),
("D4","Decision","Next adapters after MSSQL (priority order)","—","M","Confirm with stakeholders by M1 demo.","Kiran",date(2026,5,29),"Open"),
]
for i,(rid,rt,item,lik,imp,mit,owner,due,status) in enumerate(RISK,2):
    rk.cell(i,1,rid).font=BB; rk.cell(i,2,rt).font=BF; rk.cell(i,3,item).font=BF
    rk.cell(i,4,lik).font=BF; rk.cell(i,5,imp).font=BF; rk.cell(i,6,mit).font=BF; rk.cell(i,7,owner).font=BF
    rk.cell(i,8,due).number_format="yyyy-mm-dd"; rk.cell(i,9,status).font=BF
    for col in range(1,10): rk.cell(i,col).border=BORDER; rk.cell(i,col).alignment=LT if col in (3,6) else C
    rk.row_dimensions[i].height=44
dvt=DataValidation(type="list",formula1='"Risk,Decision,Issue"',allow_blank=True); dvt.add(f"B2:B{len(RISK)+1}"); rk.add_data_validation(dvt)
dvst=DataValidation(type="list",formula1='"Open,Mitigating,Closed,Escalated"',allow_blank=True); dvst.add(f"I2:I{len(RISK)+1}"); rk.add_data_validation(dvst)
for st,col in {"Open":AMBER,"Mitigating":BLUE,"Closed":GREEN,"Escalated":RED}.items():
    rk.conditional_formatting.add(f"I2:I{len(RISK)+1}",CellIsRule(operator="equal",formula=[f'"{st}"'],fill=PatternFill("solid",fgColor=col),font=Font(color=WHITE,bold=True,size=10)))
rk.conditional_formatting.add(f"E2:E{len(RISK)+1}",CellIsRule(operator="equal",formula=['"H"'],fill=PatternFill("solid",fgColor=AMBER),font=Font(color=WHITE,bold=True,size=10)))

wb.save(OUT)
print("OK ::",OUT)
print(f"Plan rows: {len(ROWS)} | Milestones: {len(MIL)} | Sprints: {len(SPR)} | Risks/Decisions: {len(RISK)}")
print(f"Dev+Stab effort excludes UAT; SP->PG marked Done (0h).")
