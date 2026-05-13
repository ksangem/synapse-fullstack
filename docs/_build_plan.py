"""Build Synapse Project Plan v1.0 — day-level WBS for 1 dev + Claude (vibe coding).

Output: Synapse_Project_Plan_v1.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import date, timedelta
from pathlib import Path

OUT = Path(__file__).parent / "Synapse_Project_Plan_v1.xlsx"

# ---------- ND palette ----------
NAVY = "002060"
DEEP = "102898"
ORANGE = "FF9933"
ALTROW = "EBF3FB"
WHITE = "FFFFFF"
GREY = "F2F2F2"
EDIT = "FFFFC0"
RAG_GREEN = "E8F5E8"
RAG_AMBER = "FFF2E5"
RAG_RED = "FDE8E8"
RAG_BLUE = "E8F4FD"
LIGHT = "EDF2FC"
BORDER_GREY = "BFBFBF"

thin = Side(style="thin", color=BORDER_GREY)
border_all = Border(top=thin, left=thin, bottom=thin, right=thin)

FONT_HDR = Font(name="Calibri", size=10, bold=True, color=WHITE)
FONT_TITLE = Font(name="Calibri", size=14, bold=True, color=NAVY)
FONT_SUB = Font(name="Calibri", size=11, bold=True, color=NAVY)
FONT_BODY = Font(name="Calibri", size=10, color="000000")
FONT_SMALL = Font(name="Calibri", size=9, color="595959")

FILL_HDR = PatternFill("solid", fgColor=NAVY)
FILL_DEEP = PatternFill("solid", fgColor=DEEP)
FILL_ORANGE = PatternFill("solid", fgColor=ORANGE)
FILL_ALT = PatternFill("solid", fgColor=ALTROW)
FILL_LIGHT = PatternFill("solid", fgColor=LIGHT)
FILL_EDIT = PatternFill("solid", fgColor=EDIT)
FILL_GREY = PatternFill("solid", fgColor=GREY)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_LEFT_C = Alignment(horizontal="left", vertical="center", wrap_text=True)

wb = Workbook()
wb.remove(wb.active)


# ============================================================
# HELPERS
# ============================================================

def write_header_row(ws, row, headers, widths=None, height=30):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = FONT_HDR
        c.fill = FILL_HDR
        c.alignment = ALIGN_CENTER
        c.border = border_all
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = height


def write_data_rows(ws, start_row, rows, alt_fill=True):
    for ri, row in enumerate(rows):
        is_alt = (ri % 2 == 0)
        for ci, val in enumerate(row, start=1):
            c = ws.cell(row=start_row + ri, column=ci, value=val)
            c.font = FONT_BODY
            c.alignment = ALIGN_LEFT
            c.border = border_all
            if alt_fill and is_alt:
                c.fill = FILL_ALT


def status_conditional(ws, status_col_letter, first_row, last_row):
    rng = f"{status_col_letter}{first_row}:{status_col_letter}{last_row}"
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="equal", formula=['"Done"'],
                   fill=PatternFill("solid", fgColor=RAG_GREEN), font=Font(name="Calibri", size=10, bold=True, color="2D6A4F"))
    )
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="equal", formula=['"In Progress"'],
                   fill=PatternFill("solid", fgColor=RAG_AMBER), font=Font(name="Calibri", size=10, bold=True, color="B45309"))
    )
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="equal", formula=['"Blocked"'],
                   fill=PatternFill("solid", fgColor=RAG_RED), font=Font(name="Calibri", size=10, bold=True, color="991B1B"))
    )
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="equal", formula=['"Not Started"'],
                   fill=PatternFill("solid", fgColor=GREY), font=Font(name="Calibri", size=10, color="595959"))
    )


def add_status_validation(ws, col_letter, first_row, last_row):
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(type="list", formula1='"Not Started,In Progress,Blocked,Done"', allow_blank=True)
    dv.add(f"{col_letter}{first_row}:{col_letter}{last_row}")
    ws.add_data_validation(dv)


def title_block(ws, title, subtitle=None, row=1):
    ws.cell(row=row, column=1, value=title).font = FONT_TITLE
    ws.row_dimensions[row].height = 24
    if subtitle:
        c = ws.cell(row=row + 1, column=1, value=subtitle)
        c.font = FONT_SMALL
        return row + 3
    return row + 2


# ============================================================
# TAB 1 — COVER & READ-ME
# ============================================================

ws = wb.create_sheet("1. Cover & Read-Me")
ws.sheet_view.showGridLines = False

# Big title block
ws.merge_cells("A1:F1")
c = ws.cell(row=1, column=1, value="SYNAPSE INTEGRATION PLATFORM — PROJECT PLAN")
c.font = Font(name="Calibri", size=18, bold=True, color=WHITE)
c.fill = FILL_DEEP
c.alignment = ALIGN_CENTER
ws.row_dimensions[1].height = 36

ws.merge_cells("A2:F2")
c = ws.cell(row=2, column=1, value="Day-level WBS for Phase 2 → Phase 4  |  Companion to Synapse_BRD_v1.docx")
c.font = Font(name="Calibri", size=11, italic=True, color=NAVY)
c.alignment = ALIGN_CENTER
ws.row_dimensions[2].height = 22

# Meta block
meta = [
    ("Version", "1.0"),
    ("Date", "2026-05-04"),
    ("Owner", "Synapse Product Team"),
    ("Classification", "Internal"),
    ("BU", "Nalashaa Digital (GLOBAL)"),
    ("Companion BRD", "Synapse_BRD_v1.docx"),
    ("Companion Charter", "Synapse_Product_Charter_v2.pdf"),
]
row = 4
for k, v in meta:
    ws.cell(row=row, column=2, value=k).font = FONT_SUB
    ws.cell(row=row, column=3, value=v).font = FONT_BODY
    row += 1

# Capacity assumption
row += 1
ws.cell(row=row, column=2, value="Capacity Assumption").font = FONT_SUB
row += 1
for line in [
    "1 full-time developer + Claude Code (vibe coding).",
    "Working hours: 6–8 productive coding hours per day (pair-programming with AI).",
    "Aggressive day-level estimates throughout — phases may compress further with strong session continuity.",
    "Estimates exclude calendar holidays; absorb interruptions in a 10% buffer per phase.",
]:
    ws.cell(row=row, column=2, value=f"• {line}").font = FONT_BODY
    row += 1

# Tab guide
row += 2
ws.cell(row=row, column=2, value="Tab Guide").font = FONT_SUB
row += 1
guide = [
    ("1. Cover & Read-Me", "This tab. Capacity, conventions, status legend, glossary."),
    ("2. Current State Snapshot", "Module status today (Built / Partial / Stub / Missing) — one-page picture."),
    ("3. WBS", "Master backlog: every story with phase, module, epic, estimate, dependencies, status."),
    ("4. Phase 2 Day Plan", "Days 1 → ~40. Goal of the day, files touched, test that proves done."),
    ("5. Phase 3 Day Plan", "Days ~41 → ~70. Scale & Governance items."),
    ("6. Phase 4 Day Plan", "Days ~71 → ~90. Polish & Distribution."),
    ("7. Milestones", "Demo and go-live checkpoints."),
    ("8. Gantt", "Visual timeline — bars per epic across the day axis."),
    ("9. Risks", "Risk register with likelihood, impact, mitigation, owner."),
    ("10. Open Decisions", "Open product / engineering decisions with decide-by day numbers."),
]
for k, v in guide:
    ws.cell(row=row, column=2, value=k).font = Font(name="Calibri", size=10, bold=True, color=NAVY)
    ws.cell(row=row, column=3, value=v).font = FONT_BODY
    row += 1

# Status legend
row += 2
ws.cell(row=row, column=2, value="Status Legend").font = FONT_SUB
row += 1
legend = [
    ("Not Started", GREY),
    ("In Progress", RAG_AMBER),
    ("Blocked", RAG_RED),
    ("Done", RAG_GREEN),
]
for label, color in legend:
    c = ws.cell(row=row, column=2, value=label)
    c.font = Font(name="Calibri", size=10, bold=True, color="000000")
    c.fill = PatternFill("solid", fgColor=color)
    c.alignment = ALIGN_CENTER
    c.border = border_all
    ws.cell(row=row, column=3, value="Editable status — drives conditional formatting throughout the workbook.").font = FONT_BODY
    row += 1

# Cell color key
row += 2
ws.cell(row=row, column=2, value="Cell Color Key").font = FONT_SUB
row += 1
key = [
    ("Editable", EDIT, "Update as you progress (Status, Notes, Actual Days)"),
    ("Calculated", GREY, "Formula-driven — do not edit"),
    ("Header", NAVY, "Tab header — locked"),
]
for label, color, note in key:
    c = ws.cell(row=row, column=2, value=label)
    fontcolor = WHITE if color == NAVY else "000000"
    c.font = Font(name="Calibri", size=10, bold=True, color=fontcolor)
    c.fill = PatternFill("solid", fgColor=color)
    c.alignment = ALIGN_CENTER
    c.border = border_all
    ws.cell(row=row, column=3, value=note).font = FONT_BODY
    row += 1

# Footer
row += 3
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
c = ws.cell(row=row, column=1, value="© 2026 Nalashaa Solutions India Pvt. Ltd. | Confidential — Internal Use Only")
c.font = FONT_SMALL
c.alignment = ALIGN_CENTER

# Column widths
ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = 28
ws.column_dimensions["C"].width = 70
ws.column_dimensions["D"].width = 12
ws.column_dimensions["E"].width = 12
ws.column_dimensions["F"].width = 12


# ============================================================
# TAB 2 — CURRENT STATE SNAPSHOT
# ============================================================

ws = wb.create_sheet("2. Current State Snapshot")
ws.sheet_view.showGridLines = False
title_block(ws, "Current State Snapshot — As of 4 May 2026",
            "Source: codebase walkthrough + PRODUCT_STATUS.md")

write_header_row(ws, 4,
                 ["#", "Module", "Status", "Notes"],
                 widths=[6, 38, 14, 70])

snapshot = [
    (1, "Jira Source — REST API (Red Gold)", "Built", "Basic Auth + JQL search, retries, custom-field parsing, normaliser."),
    (2, "Jira Source — Browser Automation (Flatiron)", "Built", "Playwright SSO/MFA, TOTP, 8h cookie session cache."),
    (3, "SharePoint Destination", "Built", "OAuth2 client credentials, Graph API site/list resolution, batched 20/req."),
    (4, "Field Mapping (Jira → SharePoint, 35 fields)", "Built", "12 standard + 23 derived (CycleTimeDays, IsOverdue, SprintNumber, etc.)."),
    (5, "Credential Encryption (AES-256-GCM)", "Built", "Store-only; no GET/decrypt endpoint yet."),
    (6, "Integration CRUD", "Built", "Zod-validated, JSONB mappings + retry policy."),
    (7, "Run Tracking", "Built", "Status pending → running → success/error, record counts."),
    (8, "Cron Scheduling (BullMQ repeatable)", "Built", "PATCH /connected/:id/schedule."),
    (9, "Delta Sync (3 modes)", "Built", "RESYNC, EXTEND_TO_TODAY, CUSTOM. Watermark-based."),
    (10, "3-Layer Dedup", "Built", "pushLog → jiraItemCache → SharePoint live query."),
    (11, "Connection Wizard (6-step)", "Partial", "Only Step 2 (Test Connection) wired. Steps 1, 3, 4, 5, 6 use mock data."),
    (12, "Mapping Canvas", "Partial", "Pure-frontend with simulated AI auto-map. Not persisted."),
    (13, "Push Page", "Partial", "Form-based; talks to /api/sharepoint/push."),
    (14, "Connector Studio", "Stub", "Test/Publish buttons disabled (StudioPage.jsx:86). No backend."),
    (15, "Health Dashboard", "Stub UI", "Beautiful UI, all mock data."),
    (16, "Integration Registry", "Stub UI", "Cards + sparklines; mock integrations.js."),
    (17, "Trading Network Console (Monitor)", "Stub UI", "Mock monitorData.js."),
    (18, "Alerts Page", "Stub UI", "Mock alerts.js. No backend dispatch."),
    (19, "Credential Vault Page", "Stub UI", "Mock credentials.js. Eye-icon UX present."),
    (20, "Master Entity Catalog", "Stub UI", "Tree + field usage on hardcoded data."),
    (21, "Connected (Integrations Dashboard)", "Stub UI", "Mock data; schedule modal works against real backend in some paths."),
    (22, "Admin (Users + Client Apps)", "Stub UI", "Mock data; no auth backend."),
    (23, "User Auth + JWT", "Missing", "users + role enum + organizations exist; never enforced."),
    (24, "RBAC Enforcement", "Missing", "No middleware checks role on routes."),
    (25, "Multi-Tenancy", "Missing", "Hardcoded org UUID."),
    (26, "Alerts System (creation + dispatch)", "Missing", "DB table + queue declared; no logic."),
    (27, "Audit Log Writer", "Missing", "DB table exists; nothing writes to it."),
    (28, "Web Scraping (Apify / Playwright actor)", "Missing", "Charter Phase 2."),
    (29, "DB / File Share / MQ Connector Types", "Missing", "Charter Phase 2."),
    (30, "AI Auto-Mapping (real LLM)", "Missing", "Stub UI only."),
    (31, "AI Help Assistant", "Missing", "Charter Phase 3."),
    (32, "Dynamics 365 / TARA / PostgreSQL / Excel / Keka connectors", "Missing", "Phase 2 expansion."),
    (33, "TFS / Holiday Tracker / Ahrefs / GSC / Google Adwords destinations", "Missing", "Phase 2-3 expansion."),
]
write_data_rows(ws, 5, snapshot)
status_conditional(ws, "C", 5, 4 + len(snapshot))

# Summary block
sr = 4 + len(snapshot) + 3
ws.cell(row=sr, column=2, value="Tally").font = FONT_SUB
sr += 1
for label, status_val in [("Built", "Built"), ("Partial", "Partial"), ("Stub UI", "Stub UI"), ("Stub", "Stub"), ("Missing", "Missing")]:
    ws.cell(row=sr, column=2, value=label).font = FONT_BODY
    c = ws.cell(row=sr, column=3, value=f'=COUNTIF(C5:C{4+len(snapshot)},"{status_val}")')
    c.font = FONT_BODY
    c.fill = FILL_GREY
    sr += 1

ws.freeze_panes = "A5"


# ============================================================
# TAB 3 — WBS (Work Breakdown Structure)
# ============================================================

ws = wb.create_sheet("3. WBS")
ws.sheet_view.showGridLines = False
title_block(ws, "Work Breakdown Structure",
            "Master backlog. Estimates in working days for 1 developer + Claude. Status drives conditional formatting.")

write_header_row(ws, 4,
                 ["ID", "Phase", "Module", "Epic", "Story", "Acceptance Criteria (summary)",
                  "Est (d)", "Dependencies", "Risk", "Status"],
                 widths=[8, 8, 22, 22, 38, 50, 8, 16, 8, 14])

# WBS items — comprehensive backlog
wbs = [
    # === FOUNDATION (Phase 2 prerequisites) ===
    ("F-01", "P2", "Foundation", "Auth & RBAC", "JWT login + auth middleware",
     "Login screen issues JWT; middleware decodes user + role; protected routes return 401 without token.", 3, "—", "Med", "Not Started"),
    ("F-02", "P2", "Foundation", "Auth & RBAC", "User CRUD + role assignment",
     "Admin can create/list/update/deactivate users; role enum enforced server-side.", 2, "F-01", "Low", "Not Started"),
    ("F-03", "P2", "Foundation", "Auth & RBAC", "Role-based route guards",
     "withRole(['admin','designer']) decorator on every state-changing route; tests cover 401/403.", 2, "F-01", "Low", "Not Started"),
    ("F-04", "P2", "Foundation", "Wiring", "Wire Dashboard page to real API",
     "Dashboard summary endpoint returns counts + sparkline data; mock removed.", 1, "F-01", "Low", "Not Started"),
    ("F-05", "P2", "Foundation", "Wiring", "Wire Registry page to real API",
     "GET /api/integrations populates cards; sparklines call /api/runs?since=7d.", 1, "F-04", "Low", "Not Started"),
    ("F-06", "P2", "Foundation", "Wiring", "Wire Connected page",
     "GET /api/connected returns groups + schedules; ConnectedPage replaces inline mock.", 1, "F-05", "Low", "Not Started"),
    ("F-07", "P2", "Foundation", "Wiring", "Wire Vault page",
     "GET /api/credentials returns metadata + mask; reveal calls /api/credentials/:id with audit log.", 1, "F-09", "Low", "Not Started"),
    ("F-08", "P2", "Foundation", "Wiring", "Wire Admin page",
     "Users + Client Apps tabs read /api/users + /api/clients; RBAC respected.", 1, "F-02", "Low", "Not Started"),
    ("F-09", "P2", "Foundation", "Vault", "Credential GET + decrypt + audit",
     "GET /api/credentials/:id decrypts and returns; reveal logged with user + IP.", 1, "F-01", "Low", "Not Started"),
    ("F-10", "P2", "Foundation", "Charts", "Add Recharts and standard chart components",
     "Recharts installed; Sparkline, BarChart, LineChart components built once and reused.", 1, "—", "Low", "Not Started"),

    # === CONNECTOR STUDIO ===
    ("CS-01", "P2", "Connector Studio", "Backend", "Connector CRUD endpoints",
     "POST/GET/PATCH/DELETE /api/connectors with versioning. app.connectors becomes active.", 2, "F-03", "Med", "Not Started"),
    ("CS-02", "P2", "Connector Studio", "Backend", "OpenAPI parser",
     "Upload Swagger/OpenAPI 3.0; backend returns endpoints + schemas.", 2, "CS-01", "Med", "Not Started"),
    ("CS-03", "P2", "Connector Studio", "Backend", "Database schema introspection",
     "Connection string → list of tables with column types via Drizzle Kit introspect.", 2, "CS-01", "Med", "Not Started"),
    ("CS-04", "P2", "Connector Studio", "Backend", "Connector test + publish",
     "POST /test runs sample call; POST /publish gated on success; new version created.", 1, "CS-02", "Low", "Not Started"),
    ("CS-05", "P2", "Connector Studio", "Frontend", "Studio UI wire-up",
     "Test Connection + Publish buttons enabled. Form posts to /api/connectors. Versions visible.", 2, "CS-04", "Low", "Not Started"),
    ("CS-06", "P2", "Connector Studio", "Frontend", "Entity modeller drag-drop",
     "Group raw fields into entities; rename labels; persist entity_definitions.", 2, "CS-05", "Med", "Not Started"),

    # === MAPPING ENGINE ===
    ("M-01", "P2", "Mapping Engine", "AI", "AI Auto-Map service (Claude API)",
     "POST /api/integrations/:id/mappings/auto-map returns suggestions + confidence ≥70% on test set.", 3, "F-10", "Med", "Not Started"),
    ("M-02", "P2", "Mapping Engine", "AI", "NL transformation generator",
     "Operator types 'convert ISO date to dd/mm/yyyy'; backend returns JS function.", 2, "M-01", "Med", "Not Started"),
    ("M-03", "P2", "Mapping Engine", "Backend", "JS sandbox (QuickJS) for transforms",
     "Transformations run in sandbox; CPU + memory limits; no host access.", 2, "M-02", "High", "Not Started"),
    ("M-04", "P2", "Mapping Engine", "Frontend", "Canvas wire-up to real API",
     "AI Auto-Map button calls real endpoint; lines colour-state machine; mappings persist.", 2, "M-01", "Low", "Not Started"),
    ("M-05", "P2", "Mapping Engine", "Frontend", "Mapping validation UI",
     "Validate button checks types + required + cycles; checklist with pass/fail.", 1, "M-04", "Low", "Not Started"),

    # === MONITORING ===
    ("MN-01", "P2", "Monitoring", "Backend", "Activate run_messages writes",
     "Every push/pull writes a row to run_messages; payload hash + status.", 1, "F-03", "Low", "Not Started"),
    ("MN-02", "P2", "Monitoring", "Backend", "Real-time stream endpoint",
     "GET /api/messages/stream (SSE) or polling; Trading Console subscribes.", 1, "MN-01", "Med", "Not Started"),
    ("MN-03", "P2", "Monitoring", "Frontend", "Trading Network Console wire-up",
     "Live table; expandable rows show payload + mapping trace; pause button freezes view.", 2, "MN-02", "Low", "Not Started"),
    ("MN-04", "P2", "Monitoring", "Backend", "Health Dashboard summary endpoints",
     "GET /api/dashboard/{summary,volume,errors-trend,client-tiers}; <500ms via materialised views.", 2, "F-03", "Med", "Not Started"),
    ("MN-05", "P2", "Monitoring", "Frontend", "Health Dashboard charts wire-up",
     "DashboardPage uses Recharts components; KPI tiles + volume + trend.", 1, "MN-04, F-10", "Low", "Not Started"),

    # === ALERTS ===
    ("AL-01", "P2", "Alerts", "Backend", "AlertService.create() + DB writes",
     "Failure paths in workers call AlertService; app.alerts populated.", 1, "F-03", "Low", "Not Started"),
    ("AL-02", "P2", "Alerts", "Backend", "alert-dispatcher worker",
     "Worker reads queue; routes to in-app + email per rule.", 1, "AL-01", "Low", "Not Started"),
    ("AL-03", "P2", "Alerts", "Backend", "Email transport (nodemailer + SMTP)",
     "SMTP env config; immediate critical, daily digest warnings.", 1, "AL-02", "Low", "Not Started"),
    ("AL-04", "P2", "Alerts", "Frontend", "AlertsPage wire-up",
     "GET /api/alerts; acknowledge/escalate buttons hit real endpoints.", 1, "AL-01", "Low", "Not Started"),
    ("AL-05", "P2", "Alerts", "Backend", "Threshold-based alerts",
     "Per-adapter or global thresholds; cron evaluates daily.", 1, "AL-04", "Low", "Not Started"),

    # === MASTER ENTITY CATALOG ===
    ("MC-01", "P2", "Master Catalog", "Backend", "entities + entity_fields tables + repo",
     "New tables, repository pattern, backfill from existing fieldMappings.", 1, "F-03", "Low", "Not Started"),
    ("MC-02", "P2", "Master Catalog", "Backend", "Entity tree + usage endpoints",
     "GET /api/entities returns tree; GET /api/entities/:id/usage returns adapter list.", 1, "MC-01", "Low", "Not Started"),
    ("MC-03", "P2", "Master Catalog", "Frontend", "CatalogPage wire-up",
     "Tree navigation; field usage bars; cross-reference links work.", 2, "MC-02", "Low", "Not Started"),

    # === HELP SYSTEM ===
    ("HS-01", "P2", "Help System", "Backend", "Markdown content + lunr.js index",
     "/help/*.md authored; build-time index generated.", 2, "—", "Low", "Not Started"),
    ("HS-02", "P2", "Help System", "Frontend", "HelpPanel wire-up + search",
     "'?' icon opens panel pre-filtered to current screen; search works.", 1, "HS-01", "Low", "Not Started"),
    ("HS-03", "P2", "Help System", "Frontend", "Tooltips + first-time walkthroughs",
     "shepherd.js walkthrough on first visit per major screen; tooltips on icons.", 2, "HS-02", "Low", "Not Started"),

    # === NEW SOURCES (Phase 2) ===
    ("NS-01", "P2", "New Sources", "Connector", "PostgreSQL source connector",
     "Schema introspection + parametrised SELECT query; entity output matches Jira format.", 2, "CS-03", "Med", "Not Started"),
    ("NS-02", "P2", "New Sources", "Connector", "Excel/CSV source connector",
     "Upload or SFTP/SharePoint location; sheet/range selection; type inference.", 1, "CS-01", "Low", "Not Started"),
    ("NS-03", "P2", "New Sources", "Connector", "Dynamics 365 source connector",
     "OData/Web API + Azure AD OAuth2; entity discovery.", 3, "CS-02", "Med", "Not Started"),
    ("NS-04", "P2", "New Sources", "Connector", "Keka source connector",
     "REST API + API Key; employee/leave entity templates.", 2, "CS-02", "Low", "Not Started"),
    ("NS-05", "P2", "New Sources", "Connector", "TARA source connector",
     "REST API or scraping fallback; project/test entity templates.", 2, "CS-02", "Med", "Not Started"),

    # === NEW DESTINATIONS (Phase 2) ===
    ("ND-01", "P2", "New Destinations", "Connector", "TFS / Azure DevOps destination",
     "REST API + PAT/OAuth; create work items; mapping template for Jira issues.", 3, "CS-02", "Med", "Not Started"),
    ("ND-02", "P2", "New Destinations", "Connector", "Holiday Tracker destination",
     "REST API or DB write per implementation discovery.", 1, "CS-01", "Low", "Not Started"),

    # === WEB SCRAPING ===
    ("WS-01", "P2", "Web Scraping", "Backend", "Apify cloud actor wrapper",
     "Configure target URL + selectors; schedule jobs; retrieve dataset.", 2, "CS-01", "Med", "Not Started"),
    ("WS-02", "P2", "Web Scraping", "Backend", "Playwright self-hosted actor",
     "Recorded flow runs; structured JSON output; screenshot on failure.", 3, "CS-01", "Med", "Not Started"),

    # === PHASE 3 ===
    ("SG-01", "P3", "Scale & Governance", "Client Apps", "Client App registration + token lifecycle",
     "External app registers; client_id/secret issued; refresh + revocation.", 2, "F-01", "Med", "Not Started"),
    ("SG-02", "P3", "Scale & Governance", "Master Catalog", "Entity merge with impact analysis",
     "Side-by-side diff; preview affected adapters; post-merge re-test flag.", 3, "MC-03", "High", "Not Started"),
    ("SG-03", "P3", "Scale & Governance", "Master Catalog", "Entity evolution / cascade",
     "Field add/rename shows downstream impact; preview-before-commit.", 2, "SG-02", "Med", "Not Started"),
    ("SG-04", "P3", "Scale & Governance", "AI Ops", "Self-healing diagnoses",
     "AI proposes fixes for common failure patterns; dry-run before apply.", 4, "M-01", "High", "Not Started"),
    ("SG-05", "P3", "Scale & Governance", "Webhooks", "Webhook support (in/out)",
     "Inbound webhook handler with HMAC signing; outbound webhook adapter type.", 2, "F-03", "Med", "Not Started"),
    ("SG-06", "P3", "Scale & Governance", "Operations", "Bulk operations",
     "Multi-select pause/resume; batch re-test after entity change.", 2, "F-05", "Low", "Not Started"),
    ("SG-07", "P3", "Scale & Governance", "Audit", "Audit middleware writes + UI",
     "Every state-changing route writes to audit_log; timeline UI filterable by user/entity/action.", 2, "F-03", "Low", "Not Started"),
    ("SG-08", "P3", "Scale & Governance", "RBAC", "Resource-level permission overrides",
     "Per-adapter Designer access for an Operator; UI to grant + revoke.", 2, "F-03", "Med", "Not Started"),
    ("SG-09", "P3", "Scale & Governance", "Help System", "AI Help Assistant",
     "Conversational chatbot with screen state + recent actions context.", 4, "HS-02, M-01", "Med", "Not Started"),
    ("SG-10", "P3", "Scale & Governance", "Connectors", "Ahrefs destination",
     "REST API + token; rank tracking entity templates.", 2, "CS-02", "Low", "Not Started"),
    ("SG-11", "P3", "Scale & Governance", "Connectors", "Google Search Console destination",
     "Service Account / OAuth; performance/coverage entities.", 2, "CS-02", "Low", "Not Started"),
    ("SG-12", "P3", "Scale & Governance", "Connectors", "Google Adwords destination",
     "Google Ads API + OAuth2 + dev token; campaign/ad-group entities.", 3, "CS-02", "Med", "Not Started"),

    # === PHASE 4 ===
    ("P4-01", "P4", "Polish", "Theming", "Theme engine (light/dark + custom logo)",
     "User toggles theme; per-org logo upload.", 1, "—", "Low", "Not Started"),
    ("P4-02", "P4", "Polish", "Mobile", "Mobile-responsive PWA",
     "Service worker, manifest, responsive break-points; monitor pages usable on tablet.", 3, "—", "Med", "Not Started"),
    ("P4-03", "P4", "Polish", "Onboarding", "First-time user onboarding wizard",
     "5-step welcome flow for first login; skippable.", 2, "F-01", "Low", "Not Started"),
    ("P4-04", "P4", "Polish", "Packaging", "Docker image + one-command deploy",
     "Single docker compose pulls image + DB + Redis; .env template.", 2, "—", "Low", "Not Started"),
    ("P4-05", "P4", "Polish", "Docs", "Architecture guide",
     "ADRs for auth, RBAC, scraping, AI mapping, encryption.", 1, "—", "Low", "Not Started"),
    ("P4-06", "P4", "Polish", "Docs", "User manual (per persona)",
     "Designer / Operator / Admin guides with screenshots.", 2, "HS-01", "Low", "Not Started"),
    ("P4-07", "P4", "Polish", "Docs", "API reference",
     "OpenAPI YAML auto-generated from routes; published as docs site.", 1, "—", "Low", "Not Started"),
    ("P4-08", "P4", "Polish", "Docs", "Handover package",
     "Bundled archive: BRD + plan + README + CLAUDE.md + runbook.", 1, "P4-05, P4-06, P4-07", "Low", "Not Started"),
    ("P4-09", "P4", "Polish", "QA", "Final QA pass + Playwright e2e suite",
     "End-to-end test per page; CI-friendly; runs <10 min.", 3, "all P2/P3", "Med", "Not Started"),
]

write_data_rows(ws, 5, wbs)
LAST_WBS = 4 + len(wbs)
status_conditional(ws, "J", 5, LAST_WBS)
add_status_validation(ws, "J", 5, LAST_WBS)
ws.freeze_panes = "A5"

# Phase totals (formula-driven)
sr = LAST_WBS + 3
ws.cell(row=sr, column=1, value="Phase Totals").font = FONT_SUB
sr += 1
write_header_row(ws, sr, ["Phase", "Items", "Estimate (d)", "Done", "Done %"], height=22)
sr += 1
for phase, label in [("P2", "Phase 2"), ("P3", "Phase 3"), ("P4", "Phase 4")]:
    ws.cell(row=sr, column=1, value=label).font = FONT_BODY
    ws.cell(row=sr, column=2, value=f'=COUNTIF(B5:B{LAST_WBS},"{phase}")').font = FONT_BODY
    ws.cell(row=sr, column=3, value=f'=SUMIF(B5:B{LAST_WBS},"{phase}",G5:G{LAST_WBS})').font = FONT_BODY
    ws.cell(row=sr, column=4, value=f'=COUNTIFS(B5:B{LAST_WBS},"{phase}",J5:J{LAST_WBS},"Done")').font = FONT_BODY
    c = ws.cell(row=sr, column=5, value=f'=IFERROR(D{sr}/B{sr},0)')
    c.font = FONT_BODY
    c.number_format = "0.0%"
    for col in range(1, 6):
        ws.cell(row=sr, column=col).fill = FILL_GREY if col != 1 else FILL_ALT
        ws.cell(row=sr, column=col).border = border_all
    sr += 1
# Grand
ws.cell(row=sr, column=1, value="GRAND TOTAL").font = FONT_SUB
ws.cell(row=sr, column=2, value=f'=COUNTA(A5:A{LAST_WBS})').font = FONT_SUB
ws.cell(row=sr, column=3, value=f'=SUM(G5:G{LAST_WBS})').font = FONT_SUB
ws.cell(row=sr, column=4, value=f'=COUNTIF(J5:J{LAST_WBS},"Done")').font = FONT_SUB
c = ws.cell(row=sr, column=5, value=f'=IFERROR(D{sr}/B{sr},0)')
c.font = FONT_SUB
c.number_format = "0.0%"
for col in range(1, 6):
    ws.cell(row=sr, column=col).fill = FILL_LIGHT
    ws.cell(row=sr, column=col).border = border_all


# ============================================================
# DAY PLAN BUILDER
# ============================================================

def write_day_plan(ws, title_text, subtitle, days_data, start_calendar_date):
    """days_data: list of (story_id, goal, files_touched, test, status)"""
    ws.sheet_view.showGridLines = False
    title_block(ws, title_text, subtitle)
    write_header_row(ws, 4,
                     ["Day #", "Calendar Date", "Story ID", "Goal of the Day",
                      "Files Touched", "Test That Proves Done", "Status"],
                     widths=[7, 14, 10, 38, 32, 32, 14])

    for i, (sid, goal, files, test, st) in enumerate(days_data):
        row = 5 + i
        is_alt = (i % 2 == 0)
        day_num = i + 1
        cal = start_calendar_date + timedelta(days=day_num - 1)
        cells = [
            day_num, cal, sid, goal, files, test, st
        ]
        for ci, val in enumerate(cells, start=1):
            c = ws.cell(row=row, column=ci, value=val)
            c.font = FONT_BODY
            c.alignment = ALIGN_LEFT
            c.border = border_all
            if is_alt:
                c.fill = FILL_ALT
        ws.cell(row=row, column=2).number_format = "yyyy-mm-dd"

    last = 4 + len(days_data)
    status_conditional(ws, "G", 5, last)
    add_status_validation(ws, "G", 5, last)
    ws.freeze_panes = "A5"

    # Summary
    sr = last + 3
    ws.cell(row=sr, column=1, value="Days Done").font = FONT_SUB
    ws.cell(row=sr, column=2, value=f'=COUNTIF(G5:G{last},"Done")').font = FONT_SUB
    ws.cell(row=sr, column=3, value=f'=COUNTIF(G5:G{last},"In Progress")').font = FONT_BODY
    ws.cell(row=sr, column=3).fill = FILL_GREY
    ws.cell(row=sr, column=4, value="In Progress").font = FONT_BODY
    sr += 1
    ws.cell(row=sr, column=1, value="Total Days").font = FONT_SUB
    ws.cell(row=sr, column=2, value=len(days_data)).font = FONT_SUB
    sr += 1
    ws.cell(row=sr, column=1, value="% Complete").font = FONT_SUB
    c = ws.cell(row=sr, column=2, value=f'=IFERROR(COUNTIF(G5:G{last},"Done")/{len(days_data)},0)')
    c.font = FONT_SUB
    c.number_format = "0.0%"


# ============================================================
# TAB 4 — PHASE 2 DAY PLAN
# ============================================================

ws = wb.create_sheet("4. Phase 2 Day Plan")

P2_START = date(2026, 5, 5)  # Day 1 = tomorrow
phase2_days = [
    # Foundation week
    ("F-01", "JWT login + auth middleware (backend)",
     "auth.routes.ts, jwtMiddleware.ts, login.tsx",
     "401 on protected route without token; 200 with valid JWT.", "Not Started"),
    ("F-01", "Login page UI + token storage (frontend)",
     "components/auth/LoginPage.jsx, services/auth.js",
     "Login form posts to /api/auth/login, stores JWT, redirects.", "Not Started"),
    ("F-02", "User CRUD endpoints + role assignment",
     "api/users.routes.ts, db/repositories/userRepo.ts",
     "Admin creates user; role enum enforced; tests cover all 4 roles.", "Not Started"),
    ("F-03", "Role-based route guards everywhere",
     "middleware/withRole.ts, all routes",
     "Admin-only route returns 403 for Operator; tests added per route.", "Not Started"),
    ("F-08", "Admin page wire-up (Users tab + Client Apps tab)",
     "components/admin/AdminPage.jsx, services/api.js",
     "AdminPage shows real users; role edit modal saves to backend.", "Not Started"),
    ("F-10", "Recharts install + Sparkline/BarChart/LineChart components",
     "components/shared/charts/*.jsx, package.json",
     "Storybook-style demo page renders all 3 chart types.", "Not Started"),
    ("F-04", "Wire Dashboard page to real summary endpoint",
     "components/dashboard/DashboardPage.jsx, api/dashboard.routes.ts",
     "KPI tiles show real counts; backend returns <500ms.", "Not Started"),
    ("F-05", "Wire Registry + sparklines",
     "components/registry/RegistryPage.jsx",
     "Cards populated from /api/integrations; sparklines plot last 7 days runs.", "Not Started"),
    ("F-06", "Wire Connected page (groupings + schedules)",
     "components/connected/ConnectedPage.jsx",
     "Cards grouped by client; schedule modal saves cron.", "Not Started"),
    ("F-09", "Credential GET + decrypt + audit reveal",
     "api/credentials.routes.ts, services/CredentialService.ts",
     "GET /api/credentials/:id decrypts; reveal logged in audit_log.", "Not Started"),
    ("F-07", "Wire Vault page (mask/reveal)",
     "components/vault/VaultPage.jsx",
     "List populates from /api/credentials; reveal calls real endpoint with timer.", "Not Started"),

    # Connector Studio
    ("CS-01", "Connector backend CRUD + versioning (part 1)",
     "api/connectors.routes.ts, db/schema.ts (connector_versions)",
     "POST/GET /api/connectors works; new version on update.", "Not Started"),
    ("CS-01", "Connector backend CRUD (part 2) + tests",
     "api/connectors.routes.ts, __tests__/connectors.test.ts",
     "Vitest suite green; RBAC enforces designer-only writes.", "Not Started"),
    ("CS-02", "OpenAPI parser service",
     "services/OpenAPIParser.ts",
     "Upload OpenAPI 3.0 → returns endpoints + schemas.", "Not Started"),
    ("CS-02", "Operation discovery + filtering UI",
     "components/studio/OperationDiscovery.jsx",
     "Designer ticks operations to expose; tags Read/Write/Both.", "Not Started"),
    ("CS-03", "Database schema introspection",
     "services/DBIntrospector.ts",
     "Postgres connection string → table list with column types.", "Not Started"),
    ("CS-04", "Connector test + publish endpoints",
     "api/connectors.routes.ts (publish), services/ConnectorTester.ts",
     "Publish gated on successful test; version increments.", "Not Started"),
    ("CS-05", "Studio UI wire-up (enable Test/Publish buttons)",
     "components/studio/StudioPage.jsx",
     "StudioPage.jsx:86 buttons enabled; round-trip works.", "Not Started"),
    ("CS-06", "Entity modeller drag-drop (part 1)",
     "components/studio/EntityModeller.jsx",
     "Designer groups raw fields into entities; labels editable.", "Not Started"),
    ("CS-06", "Entity modeller (part 2) + persist entity_definitions",
     "db/schema.ts (entity_definitions), components/studio/EntityModeller.jsx",
     "Save persists; reload restores layout.", "Not Started"),

    # Master Catalog
    ("MC-01", "Master Catalog tables + repo",
     "db/schema.ts (entities, entity_fields), db/repositories/entityRepo.ts",
     "Tables created via Drizzle migration; repo CRUD tested.", "Not Started"),
    ("MC-02", "Entity tree + usage endpoints",
     "api/entities.routes.ts",
     "GET /api/entities returns tree; usage count joins integrations.fieldMappings.", "Not Started"),
    ("MC-03", "CatalogPage wire-up (part 1) — tree nav",
     "components/catalog/CatalogPage.jsx",
     "Tree from real API; clicking entity loads fields panel.", "Not Started"),
    ("MC-03", "CatalogPage (part 2) — field usage bars + cross-ref",
     "components/catalog/CatalogPage.jsx, components/shared/charts/UsageBar.jsx",
     "Field usage statistics displayed; click to see adapters.", "Not Started"),

    # AI Mapping
    ("M-01", "Anthropic SDK + AI Auto-Map service (part 1)",
     "services/MappingAIService.ts, package.json",
     "POST /api/integrations/:id/mappings/auto-map returns suggestions.", "Not Started"),
    ("M-01", "AI Auto-Map (part 2) — confidence scoring + tests",
     "services/MappingAIService.ts, __tests__/mappingAI.test.ts",
     "Test set hits ≥70% accuracy on known Jira→SP mappings.", "Not Started"),
    ("M-01", "AI Auto-Map (part 3) — caching + cost guardrails",
     "services/MappingAIService.ts (cache layer)",
     "Repeated identical request hits cache; daily cost cap enforced.", "Not Started"),
    ("M-02", "NL transformation generator",
     "services/NLTransformService.ts, api/integrations.routes.ts",
     "Operator NL input → JS function returned; sample tests pass.", "Not Started"),
    ("M-03", "QuickJS sandbox (part 1) — runtime + limits",
     "services/SandboxService.ts, package.json",
     "Sandbox spawns; CPU/memory limits enforced; no host fs access.", "Not Started"),
    ("M-03", "QuickJS sandbox (part 2) — wire into push pipeline",
     "services/SharePointPushService.ts (transformations)",
     "Transformations applied per mapping line during real push.", "Not Started"),
    ("M-04", "Canvas wire-up to real AI + persistence",
     "components/canvas/CanvasPage.jsx, services/api.js",
     "AI Auto-Map button calls real endpoint; mapping saves to integrations.fieldMappings.", "Not Started"),
    ("M-05", "Mapping validation UI + checklist",
     "components/canvas/CanvasPage.jsx (validate panel)",
     "Validate button: types compatible, required mapped, no cycles.", "Not Started"),

    # Monitoring
    ("MN-01", "Activate run_messages writes",
     "workers/syncWorker.ts, services/SharePointPushService.ts",
     "Each push/pull writes a row; payload hash stored.", "Not Started"),
    ("MN-02", "Real-time stream endpoint (SSE)",
     "api/messages.routes.ts (stream)",
     "SSE feed sends new rows in <1s; client reconnects on drop.", "Not Started"),
    ("MN-03", "Trading Network Console wire-up",
     "components/monitor/MonitorPage.jsx",
     "Live table + expand rows + payload + mapping trace + pause.", "Not Started"),
    ("MN-04", "Health Dashboard summary endpoints",
     "api/dashboard.routes.ts, db/views.sql (materialised)",
     "All 4 endpoints return <500ms.", "Not Started"),
    ("MN-05", "Health Dashboard charts wire-up",
     "components/dashboard/DashboardPage.jsx",
     "Volume bar + trend line + client tier ribbon all live.", "Not Started"),

    # Alerts
    ("AL-01", "AlertService.create() + DB writes from worker failure paths",
     "services/AlertService.ts, workers/integrationRunner.ts",
     "Failing run inserts alert row with severity + integration ref.", "Not Started"),
    ("AL-02", "alert-dispatcher worker logic",
     "workers/alertDispatcher.ts",
     "Worker reads queue; routes to in-app + email per user prefs.", "Not Started"),
    ("AL-03", "Email transport (nodemailer + SMTP)",
     "services/EmailService.ts, .env",
     "Critical alert delivers email; daily digest job scheduled.", "Not Started"),
    ("AL-04", "AlertsPage wire-up + acknowledge/escalate",
     "components/alerts/AlertsPage.jsx",
     "Acknowledge button calls real API; alert moves to acknowledged state.", "Not Started"),
    ("AL-05", "Threshold-based custom alerts",
     "services/ThresholdService.ts, daily cron job",
     "Threshold breach generates alert; admin UI to configure.", "Not Started"),

    # Help System
    ("HS-01", "Help content authoring + lunr index (part 1)",
     "docs/help/*.md, scripts/build-help-index.js",
     "10+ articles authored; index built at build time.", "Not Started"),
    ("HS-01", "Help content (part 2) — per-persona guides + screenshots",
     "docs/help/*.md, public/help/screenshots/*",
     "Designer/Operator/Admin guides complete; screenshots embedded.", "Not Started"),
    ("HS-02", "HelpPanel wire-up + search",
     "components/layout/HelpPanel.jsx",
     "'?' icon opens panel; search returns ranked results.", "Not Started"),
    ("HS-03", "Tooltips + first-time walkthroughs",
     "components/shared/Tooltip.jsx, walkthroughs/*.js (shepherd.js)",
     "Tooltip on every icon; walkthrough on first visit per major screen.", "Not Started"),
    ("HS-03", "Walkthrough polish + dismissable + re-accessible",
     "components/layout/HelpMenu.jsx",
     "User can re-launch walkthrough from Help menu.", "Not Started"),

    # New Sources
    ("NS-01", "PostgreSQL source connector",
     "integrations/postgres/*, services/PostgresExtractor.ts",
     "Schema introspect + parametrised SELECT; entities match Jira format.", "Not Started"),
    ("NS-01", "PostgreSQL pagination + delta sync support",
     "services/PostgresExtractor.ts (cursor-based)",
     "Large tables paginate; delta query via timestamp watermark.", "Not Started"),
    ("NS-02", "Excel/CSV source connector",
     "integrations/excel/*, services/ExcelParser.ts",
     "Upload XLSX or read from SharePoint path; sheet/range selection.", "Not Started"),
    ("NS-03", "Dynamics 365 source connector (part 1) — auth + discovery",
     "integrations/dynamics/*, services/DynamicsClient.ts",
     "Azure AD OAuth2; entity discovery via OData $metadata.", "Not Started"),
    ("NS-03", "Dynamics 365 (part 2) — entity templates",
     "integrations/dynamics/templates/*",
     "Account/Contact/Lead templates with default field labels.", "Not Started"),
    ("NS-03", "Dynamics 365 (part 3) — testing + delta sync",
     "integrations/dynamics/*, __tests__/dynamics.test.ts",
     "Delta sync via modifiedon timestamp; test against sandbox.", "Not Started"),
    ("NS-04", "Keka source connector (part 1)",
     "integrations/keka/*",
     "REST API + API Key; employee/leave templates.", "Not Started"),
    ("NS-04", "Keka source connector (part 2) — testing + docs",
     "integrations/keka/*, docs/help/keka-connector.md",
     "Sandbox tested; help article published.", "Not Started"),
    ("NS-05", "TARA source connector (part 1)",
     "integrations/tara/*",
     "REST API or scraping fallback discovered; project/test entities.", "Not Started"),
    ("NS-05", "TARA source connector (part 2) — finalize",
     "integrations/tara/*",
     "Templates + tests; help article.", "Not Started"),

    # New Destinations (Phase 2 subset)
    ("ND-01", "TFS / Azure DevOps destination (part 1)",
     "integrations/tfs/*, services/TFSPushService.ts",
     "REST API + PAT; create work items in dev project.", "Not Started"),
    ("ND-01", "TFS (part 2) — mapping template + dedup",
     "integrations/tfs/templates/*",
     "Jira issue → Work Item mapping template; dedup on Title.", "Not Started"),
    ("ND-01", "TFS (part 3) — tests + docs",
     "__tests__/tfs.test.ts, docs/help/tfs-destination.md",
     "Vitest green; help article live.", "Not Started"),
    ("ND-02", "Holiday Tracker destination",
     "integrations/holiday-tracker/*",
     "API or DB write; mapping template; test against staging.", "Not Started"),

    # Web Scraping
    ("WS-01", "Apify cloud actor wrapper (part 1)",
     "integrations/scrape/apify/*, services/ApifyClient.ts",
     "Configure target URL + selectors; trigger run; retrieve dataset.", "Not Started"),
    ("WS-01", "Apify (part 2) — scheduling + monitoring",
     "services/ApifyClient.ts, workers/scrapeRunner.ts",
     "Apify schedule integrated with BullMQ; failure alert dispatched.", "Not Started"),
    ("WS-02", "Playwright actor (part 1) — runtime",
     "integrations/scrape/playwright/*",
     "Recorded flow runs in worker; structured JSON returned.", "Not Started"),
    ("WS-02", "Playwright actor (part 2) — screenshot diff on failure",
     "integrations/scrape/playwright/*",
     "Failure captures screenshot; alert includes diff image.", "Not Started"),
    ("WS-02", "Playwright actor (part 3) — anti-detection basics + polish",
     "integrations/scrape/playwright/*",
     "Rotating UA; configurable delays; spider/bot icon in registry.", "Not Started"),
]

write_day_plan(ws, "Phase 2 — Intelligence & Monitoring + Multi-Source",
               f"Days 1 → {len(phase2_days)}. Capacity: 1 dev + Claude. Aggressive vibe-coded pace.",
               phase2_days, P2_START)
P2_LAST_DAY = len(phase2_days)


# ============================================================
# TAB 5 — PHASE 3 DAY PLAN
# ============================================================

ws = wb.create_sheet("5. Phase 3 Day Plan")

P3_START = P2_START + timedelta(days=P2_LAST_DAY)

phase3_days = [
    ("SG-01", "Client App Registration backend + token issuance",
     "api/clients.routes.ts, db/schema.ts (client_apps)",
     "External app registers; client_id + secret returned; stored hashed.", "Not Started"),
    ("SG-01", "Client App lifecycle (refresh + revoke + per-client dashboard)",
     "components/admin/ClientApps.jsx, api/clients.routes.ts",
     "Refresh + revoke endpoints; admin sees per-client volume + tier.", "Not Started"),
    ("SG-02", "Entity merge — backend impact analysis",
     "services/EntityMergeService.ts, api/entities.routes.ts",
     "POST /:id/merge dry-run returns affected adapters list.", "Not Started"),
    ("SG-02", "Entity merge — UI side-by-side diff",
     "components/catalog/MergeModal.jsx",
     "Designer selects 2 entities; diff shown; pick fields to keep.", "Not Started"),
    ("SG-02", "Entity merge — commit + post-merge re-test flag",
     "services/EntityMergeService.ts (commit), workers/integrationRunner.ts",
     "On commit, affected integrations flagged; banner on Registry cards.", "Not Started"),
    ("SG-03", "Entity evolution — add/rename field with cascade preview",
     "services/EntityEvolutionService.ts, components/catalog/AddFieldModal.jsx",
     "Designer adds field; sees downstream mapping impact before commit.", "Not Started"),
    ("SG-03", "Entity evolution — preview-before-commit + rollback",
     "services/EntityEvolutionService.ts (rollback)",
     "Mistakes can be reverted within 1 hour window.", "Not Started"),
    ("SG-04", "Self-healing service (part 1) — diagnose common failures",
     "services/SelfHealingService.ts",
     "Failure pattern catalog; LLM-driven root-cause classifier.", "Not Started"),
    ("SG-04", "Self-healing (part 2) — propose fix + dry-run",
     "services/SelfHealingService.ts (propose)",
     "Auto-suggested fixes shown to admin; dry-run validates before apply.", "Not Started"),
    ("SG-04", "Self-healing (part 3) — auto-apply opt-in + audit",
     "services/SelfHealingService.ts (apply), middleware/audit.ts",
     "Admin opts adapter into auto-apply; every fix logged.", "Not Started"),
    ("SG-04", "Self-healing (part 4) — UI banners + reporting",
     "components/dashboard/HealingBanner.jsx, components/dashboard/HealingReport.jsx",
     "Banner on dashboard; weekly self-healing report.", "Not Started"),
    ("SG-05", "Webhook support — outbound adapter",
     "integrations/webhook/*, services/WebhookPushService.ts",
     "Outbound webhook with HMAC sign; delivery retry policy.", "Not Started"),
    ("SG-05", "Webhook support — inbound handler + ingest",
     "api/webhooks.routes.ts (inbound), services/WebhookIngestService.ts",
     "Inbound webhook validates HMAC; pushes to integration as if API event.", "Not Started"),
    ("SG-06", "Bulk operations — multi-select + actions",
     "components/registry/RegistryPage.jsx (selection), api/integrations.routes.ts (bulk)",
     "Select multiple → Pause/Resume/Re-test actions; confirm modal.", "Not Started"),
    ("SG-06", "Bulk operations — batch re-test after entity change",
     "services/BulkOpsService.ts",
     "Triggered from entity merge/evolution flow; progress shown.", "Not Started"),
    ("SG-07", "Audit middleware writes — every state-changing route",
     "middleware/audit.ts, all routes",
     "Each POST/PATCH/DELETE writes audit row with diff + user + IP.", "Not Started"),
    ("SG-07", "Audit log UI timeline + filters",
     "components/admin/AuditTimeline.jsx",
     "Timeline view; filter by user, entity, action; export CSV.", "Not Started"),
    ("SG-08", "Resource-level RBAC — backend",
     "db/schema.ts (user_overrides), middleware/withRole.ts",
     "Per-adapter Designer access for an Operator on one connector only.", "Not Started"),
    ("SG-08", "Resource-level RBAC — admin UI",
     "components/admin/PermissionOverrides.jsx",
     "Admin grants/revokes resource-level permissions.", "Not Started"),
    ("SG-09", "AI Help Assistant (part 1) — chat backend",
     "services/HelpAssistantService.ts, api/help.routes.ts",
     "POST /api/help/ask; LLM with help library + screen state context.", "Not Started"),
    ("SG-09", "AI Help Assistant (part 2) — UI in HelpPanel",
     "components/layout/HelpPanel.jsx (chat tab)",
     "Chat tab in help panel; conversational; suggests next steps.", "Not Started"),
    ("SG-09", "AI Help Assistant (part 3) — context awareness",
     "services/HelpAssistantService.ts (context loader)",
     "Assistant sees current screen + recent actions; suggests context-aware next steps.", "Not Started"),
    ("SG-09", "AI Help Assistant (part 4) — guardrails + cost cap",
     "services/HelpAssistantService.ts (rate limit + token cap)",
     "Per-user daily token cap; rate limit; safety filter on output.", "Not Started"),
    ("SG-10", "Ahrefs destination connector",
     "integrations/ahrefs/*, services/AhrefsPushService.ts",
     "REST API + token; rank tracking entity templates; tests.", "Not Started"),
    ("SG-11", "Google Search Console destination",
     "integrations/gsc/*, services/GSCPushService.ts",
     "Service account / OAuth; performance + coverage entities.", "Not Started"),
    ("SG-11", "GSC — testing + docs",
     "__tests__/gsc.test.ts, docs/help/gsc-destination.md",
     "Vitest green; help article live.", "Not Started"),
    ("SG-12", "Google Adwords destination (part 1) — auth + dev token",
     "integrations/google-ads/*",
     "OAuth2 + developer token; refresh token storage.", "Not Started"),
    ("SG-12", "Google Adwords destination (part 2) — campaigns/ad-groups",
     "integrations/google-ads/templates/*",
     "Entity templates; create + update via Google Ads API.", "Not Started"),
    ("SG-12", "Google Adwords destination (part 3) — testing + go-live",
     "__tests__/google-ads.test.ts",
     "Sandbox account verified; help article live.", "Not Started"),
]

write_day_plan(ws, "Phase 3 — Scale & Governance",
               f"Days {P2_LAST_DAY+1} → {P2_LAST_DAY+len(phase3_days)}. Multi-tenant + governance + Phase-3 destinations.",
               phase3_days, P3_START)
P3_LAST_DAY = P2_LAST_DAY + len(phase3_days)


# ============================================================
# TAB 6 — PHASE 4 DAY PLAN
# ============================================================

ws = wb.create_sheet("6. Phase 4 Day Plan")

P4_START = P3_START + timedelta(days=len(phase3_days))

phase4_days = [
    ("P4-01", "Theme engine (light/dark + custom logo upload)",
     "contexts/ThemeContext.jsx, components/admin/BrandingTab.jsx",
     "Toggle persists per user; logo upload validated + stored.", "Not Started"),
    ("P4-02", "Mobile PWA — manifest + service worker",
     "public/manifest.json, src/serviceWorker.js",
     "Lighthouse PWA score >90; installable on Android.", "Not Started"),
    ("P4-02", "Mobile PWA — responsive break-points",
     "src/styles.css (media queries)",
     "Dashboard + Monitor + Alerts usable on tablet (768px) and phone (375px).", "Not Started"),
    ("P4-02", "Mobile PWA — offline fallback for critical pages",
     "src/serviceWorker.js (cache strategies)",
     "Last-seen Health Dashboard renders offline.", "Not Started"),
    ("P4-03", "Onboarding wizard (part 1) — 5-step welcome",
     "components/onboarding/OnboardingWizard.jsx",
     "First-login users walk through 5 screens; skip + replay.", "Not Started"),
    ("P4-03", "Onboarding wizard (part 2) — sample data + sandbox",
     "components/onboarding/SampleData.jsx",
     "User can spin up demo Jira→SP integration in <2 minutes.", "Not Started"),
    ("P4-04", "Docker image build + compose",
     "Dockerfile, docker-compose.prod.yml",
     "Single image runs FE + BE; compose includes Postgres + Redis.", "Not Started"),
    ("P4-04", "One-command deploy script + .env.template",
     "scripts/deploy.sh, .env.template",
     "Fresh server: docker compose up reaches healthy state.", "Not Started"),
    ("P4-05", "Architecture guide + ADRs",
     "docs/architecture.md, docs/adr/*.md",
     "ADRs for auth, RBAC, sandboxing, AI mapping, encryption.", "Not Started"),
    ("P4-06", "User manual — Operator guide",
     "docs/user-manual/operator.md (+ screenshots)",
     "Step-by-step with screenshots; reviewed by 1 BA.", "Not Started"),
    ("P4-06", "User manual — Designer + Admin guides",
     "docs/user-manual/{designer,admin}.md",
     "All three guides published; cross-linked from Help panel.", "Not Started"),
    ("P4-07", "API reference (OpenAPI auto-gen)",
     "scripts/generate-openapi.js, docs/api/*",
     "OpenAPI YAML generated from routes; published as Redoc site.", "Not Started"),
    ("P4-08", "Handover package (zip everything)",
     "scripts/handover-pack.sh",
     "Single archive: BRD + plan + manuals + ADRs + runbook.", "Not Started"),
    ("P4-09", "Final QA — Playwright e2e suite (part 1)",
     "e2e/*.spec.ts (Wizard, Mapping, Push)",
     "5 critical flows green in <10 min CI run.", "Not Started"),
    ("P4-09", "Final QA — Playwright e2e (part 2 — Admin + Alerts)",
     "e2e/admin.spec.ts, e2e/alerts.spec.ts",
     "User mgmt + alert flow tests green.", "Not Started"),
    ("P4-09", "Final QA — Playwright e2e (part 3 — multi-connector regression)",
     "e2e/multi-connector.spec.ts",
     "Dynamics, PostgreSQL, Excel, Keka, TARA, TFS, GSC sample paths green.", "Not Started"),
    ("P4-09", "Final go-live demo + retrospective",
     "—",
     "Demo to stakeholders; lessons-learned log captured.", "Not Started"),
]

write_day_plan(ws, "Phase 4 — Polish & Distribution",
               f"Days {P3_LAST_DAY+1} → {P3_LAST_DAY+len(phase4_days)}. Theming + PWA + onboarding + Docker + docs + final QA.",
               phase4_days, P4_START)
P4_LAST_DAY = P3_LAST_DAY + len(phase4_days)


# ============================================================
# TAB 7 — MILESTONES
# ============================================================

ws = wb.create_sheet("7. Milestones")
ws.sheet_view.showGridLines = False
title_block(ws, "Milestones & Demos",
            "Key checkpoint dates calibrated to the day plans. Update Status as each is reached.")

write_header_row(ws, 4,
                 ["#", "Milestone", "Phase", "Day #", "Calendar Date", "Outcome / Demo Script", "Status"],
                 widths=[6, 30, 8, 8, 14, 60, 14])

milestones = [
    (1, "Auth + RBAC live", "P2", 5, "Login screen works; protected routes enforce role; audit log captures auth events."),
    (2, "Frontend wired (Dashboard, Registry, Vault, Admin)", "P2", 11, "End of Foundation week — no more mock data on these pages."),
    (3, "Connector Studio enabled", "P2", 20, "Designer can publish a new REST connector end-to-end (Test → Publish → appears in Registry)."),
    (4, "Master Catalog read-mode live", "P2", 24, "CatalogPage shows real entities + field usage; cross-references work."),
    (5, "AI Auto-Map demo", "P2", 32, "Wizard step 4 / Canvas: AI suggestions, NL transforms, sandbox transformations applied to a real push."),
    (6, "Trading Console + Health Dashboard live", "P2", 37, "Real-time message stream visible; KPI tiles + charts populated; pause-and-inspect demo."),
    (7, "Alerts dispatching (in-app + email)", "P2", 42, "Manufactured failure → email to Operator + Designer within 5 minutes."),
    (8, "Multi-source demo (PostgreSQL + Excel)", "P2", 50, "Two new sources demoed; same wizard flow as Jira."),
    (9, "Dynamics + Keka + TARA online", "P2", 60, "Three more sources live; coverage matrix updated."),
    (10, "TFS + Holiday Tracker destinations", "P2", 65, "Two destinations beyond SharePoint live; Jira→TFS demo."),
    (11, "Web Scraping (Apify + Playwright)", "P2", P2_LAST_DAY, "End of Phase 2. One Apify and one Playwright scraping connector running on a schedule."),
    (12, "Client App Registration + bulk ops", "P3", P2_LAST_DAY + 6, "External app registers, gets token, calls API; bulk pause demoed."),
    (13, "Entity merge + evolution", "P3", P2_LAST_DAY + 12, "Two overlapping entities merged with impact preview; field added with cascade."),
    (14, "Self-healing live", "P3", P2_LAST_DAY + 18, "Common failure auto-fixed in dry-run; admin opts adapter into auto-apply."),
    (15, "AI Help Assistant", "P3", P2_LAST_DAY + 26, "Conversational help with screen-state context; demo end-to-end."),
    (16, "Phase 3 destinations (Ahrefs + GSC + Adwords)", "P3", P3_LAST_DAY, "End of Phase 3. All charter destinations on the board."),
    (17, "Theme engine + PWA", "P4", P3_LAST_DAY + 4, "Light/dark toggle; install on phone; offline Health Dashboard."),
    (18, "Onboarding wizard live", "P4", P3_LAST_DAY + 6, "First-login flow demoed; new user gets to a sample integration in <2 min."),
    (19, "Docker one-command deploy", "P4", P3_LAST_DAY + 8, "Fresh VM → docker compose up → app reachable + healthy."),
    (20, "Handover package + docs complete", "P4", P3_LAST_DAY + 13, "Architecture guide + 3 user manuals + API ref + handover zip."),
    (21, "Production handover demo", "P4", P4_LAST_DAY, "Final stakeholder demo; project plan signed off."),
]

for i, (num, name, phase, day, outcome) in enumerate(milestones):
    row = 5 + i
    cal = P2_START + timedelta(days=day - 1)
    cells = [num, name, phase, day, cal, outcome, "Not Started"]
    for ci, val in enumerate(cells, start=1):
        c = ws.cell(row=row, column=ci, value=val)
        c.font = FONT_BODY
        c.alignment = ALIGN_LEFT
        c.border = border_all
        if i % 2 == 0:
            c.fill = FILL_ALT
    ws.cell(row=row, column=5).number_format = "yyyy-mm-dd"

last_ms = 4 + len(milestones)
status_conditional(ws, "G", 5, last_ms)
add_status_validation(ws, "G", 5, last_ms)
ws.freeze_panes = "A5"


# ============================================================
# TAB 8 — GANTT
# ============================================================

ws = wb.create_sheet("8. Gantt")
ws.sheet_view.showGridLines = False
title_block(ws, "Gantt — Epic Timeline",
            f"Total span: Day 1 → Day {P4_LAST_DAY} ({P2_START} → {P2_START + timedelta(days=P4_LAST_DAY-1)}). Bars filled in NAVY; phase bands in DEEP/ORANGE.")

# Header: Epic | Phase | Start | End | Duration | day columns 1..N
TOTAL_DAYS = P4_LAST_DAY
header = ["Epic", "Phase", "Start Day", "End Day", "Duration"] + [str(d) for d in range(1, TOTAL_DAYS + 1)]
write_header_row(ws, 4, header, widths=[24, 7, 8, 8, 8] + [3] * TOTAL_DAYS, height=22)
ws.row_dimensions[4].height = 24

# Calculate Gantt bars from day plan offsets
epics = [
    # (Epic name, Phase, start_day, end_day)
    ("Foundation: Auth + Wiring", "P2", 1, 11),
    ("Connector Studio", "P2", 12, 20),
    ("Master Entity Catalog (read)", "P2", 21, 24),
    ("Mapping Engine + AI Auto-Map", "P2", 25, 32),
    ("Monitoring (Console + Dashboard)", "P2", 33, 37),
    ("Alerts System", "P2", 38, 42),
    ("Help System", "P2", 43, 47),
    ("New Sources (PG/Excel/D365/Keka/TARA)", "P2", 48, 60),
    ("New Destinations (TFS + Holiday)", "P2", 61, 65),
    ("Web Scraping (Apify + Playwright)", "P2", 66, P2_LAST_DAY),
    ("Client Apps + Lifecycle", "P3", P2_LAST_DAY + 1, P2_LAST_DAY + 2),
    ("Entity Merge + Evolution", "P3", P2_LAST_DAY + 3, P2_LAST_DAY + 7),
    ("Self-Healing", "P3", P2_LAST_DAY + 8, P2_LAST_DAY + 11),
    ("Webhook Support + Bulk Ops", "P3", P2_LAST_DAY + 12, P2_LAST_DAY + 15),
    ("Audit + Resource RBAC", "P3", P2_LAST_DAY + 16, P2_LAST_DAY + 19),
    ("AI Help Assistant", "P3", P2_LAST_DAY + 20, P2_LAST_DAY + 23),
    ("Phase 3 Destinations (Ahrefs/GSC/Adwords)", "P3", P2_LAST_DAY + 24, P3_LAST_DAY),
    ("Theme + PWA", "P4", P3_LAST_DAY + 1, P3_LAST_DAY + 4),
    ("Onboarding + Packaging", "P4", P3_LAST_DAY + 5, P3_LAST_DAY + 8),
    ("Documentation", "P4", P3_LAST_DAY + 9, P3_LAST_DAY + 13),
    ("Final QA + Demo", "P4", P3_LAST_DAY + 14, P4_LAST_DAY),
]

phase_color = {
    "P2": "1E5BB8",  # blue
    "P3": "C77600",  # orange-brown
    "P4": "2D6A4F",  # green
}

for i, (epic, phase, sd, ed) in enumerate(epics):
    row = 5 + i
    is_alt = (i % 2 == 0)
    # Static cols
    statics = [epic, phase, sd, ed, ed - sd + 1]
    for ci, val in enumerate(statics, start=1):
        c = ws.cell(row=row, column=ci, value=val)
        c.font = FONT_BODY
        c.alignment = ALIGN_LEFT_C
        c.border = border_all
        if is_alt:
            c.fill = FILL_ALT
    # Day cells
    bar_fill = PatternFill("solid", fgColor=phase_color[phase])
    for d in range(1, TOTAL_DAYS + 1):
        c = ws.cell(row=row, column=5 + d, value="")
        c.border = border_all
        if sd <= d <= ed:
            c.fill = bar_fill
        elif is_alt:
            c.fill = FILL_ALT

last_g = 4 + len(epics)
ws.freeze_panes = "F5"

# Legend
sr = last_g + 3
ws.cell(row=sr, column=1, value="Legend").font = FONT_SUB
sr += 1
for label, hexcol in [("Phase 2", "1E5BB8"), ("Phase 3", "C77600"), ("Phase 4", "2D6A4F")]:
    c = ws.cell(row=sr, column=1, value=label)
    c.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=hexcol)
    c.alignment = ALIGN_CENTER
    c.border = border_all
    sr += 1


# ============================================================
# TAB 9 — RISKS
# ============================================================

ws = wb.create_sheet("9. Risks")
ws.sheet_view.showGridLines = False
title_block(ws, "Risk Register",
            "Charter risks + code-level risks discovered during walkthrough. Update Status weekly.")

write_header_row(ws, 4,
                 ["#", "Risk", "Likelihood", "Impact", "Mitigation", "Owner", "Status"],
                 widths=[5, 38, 12, 10, 50, 14, 14])

risks = [
    (1, "External API access not obtainable for some destinations", "Medium", "High",
     "Three-tier fallback: browser passthrough → Apify cloud → Playwright self-hosted. Scraping connectors flagged with bot icon and monitored.", "Designer", "Not Started"),
    (2, "Web scraping connectors break on UI changes", "High", "Medium",
     "Screenshot-diff alerting on scrape failures; designer notified immediately with diff image.", "Designer", "Not Started"),
    (3, "Scope creep expands beyond planned phases", "High", "High",
     "Lock scope per phase; change-request process; MVP-first discipline; milestones reviewed weekly.", "Product Owner", "Not Started"),
    (4, "AI Auto-Map produces poor suggestions", "Medium", "Medium",
     "Always require human confirmation; log decisions for tuning; graceful fallback to manual mapping. Target ≥70% accuracy.", "Designer", "Not Started"),
    (5, "Non-technical users find mapping too complex", "Medium", "High",
     "Extensive UX testing; progressive disclosure; task escalation to Designer; first-time walkthroughs.", "PM", "Not Started"),
    (6, "Credential vault security vulnerability", "Low", "Critical",
     "AES-256-GCM; security review before any external exposure; no raw credential display anywhere; reveal events audited.", "Admin", "Not Started"),
    (7, "10 of 11 frontend pages on mock data — silent regressions when wired", "High", "Medium",
     "End-to-end Playwright test added at the time of wiring each page (test-first).", "Developer", "Not Started"),
    (8, "No auth enforcement today — accidental public exposure if deployed", "Medium", "Critical",
     "Lock to localhost / VPN until F-01/F-02/F-03 ship; no production deploy until RBAC live.", "Admin", "Not Started"),
    (9, "Single-developer bus factor + vibe-coding velocity dependence", "Medium", "High",
     "Strong PRODUCT_STATUS.md + BRD + day-level plan keep the project transferable; CLAUDE.md captures session conventions.", "PM", "Not Started"),
    (10, "AI mapping cost overruns (Anthropic API)", "Medium", "Medium",
     "Daily cost cap per environment; cache layer; configurable rate limit; monitor in dashboard.", "Developer", "Not Started"),
    (11, "JS sandbox escape risk in transformations", "Low", "Critical",
     "QuickJS in worker thread; strict CPU/memory limits; no host fs/network access from sandbox; security review.", "Developer", "Not Started"),
    (12, "Long-running queue jobs starve other work", "Medium", "Medium",
     "Per-queue concurrency limits; separate queue for scraping; BullMQ priority levels for critical alerts.", "Developer", "Not Started"),
    (13, "Schema migration breaks running adapters", "Medium", "High",
     "Drizzle migrations versioned; pre-deploy dry-run on shadow DB; rollback plan documented per migration.", "Developer", "Not Started"),
    (14, "Email transport delivery issues (alerts ignored)", "Low", "High",
     "Health probe sends test email weekly; SMTP fallback to alternate provider; in-app bell as backup channel.", "Admin", "Not Started"),
    (15, "Tech-stack lock-in to React 19 / Express 5 / Drizzle", "Low", "Low",
     "Documented in BRD §11.2. Acceptable trade-off for velocity; modular service layer minimises lock-in cost.", "Architect", "Not Started"),
]

for i, (num, risk, lik, imp, mit, owner, status) in enumerate(risks):
    row = 5 + i
    is_alt = (i % 2 == 0)
    cells = [num, risk, lik, imp, mit, owner, status]
    for ci, val in enumerate(cells, start=1):
        c = ws.cell(row=row, column=ci, value=val)
        c.font = FONT_BODY
        c.alignment = ALIGN_LEFT
        c.border = border_all
        if is_alt:
            c.fill = FILL_ALT

last_r = 4 + len(risks)
status_conditional(ws, "G", 5, last_r)
add_status_validation(ws, "G", 5, last_r)

# Likelihood/Impact conditional formatting
for col in ("C", "D"):
    rng = f"{col}5:{col}{last_r}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"High"'], fill=PatternFill("solid", fgColor=RAG_RED)))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Medium"'], fill=PatternFill("solid", fgColor=RAG_AMBER)))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Low"'], fill=PatternFill("solid", fgColor=RAG_GREEN)))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Critical"'], fill=PatternFill("solid", fgColor=RAG_RED), font=Font(bold=True, color="FFFFFF")))

ws.freeze_panes = "A5"


# ============================================================
# TAB 10 — OPEN DECISIONS
# ============================================================

ws = wb.create_sheet("10. Open Decisions")
ws.sheet_view.showGridLines = False
title_block(ws, "Open Decisions",
            "Charter open decisions + new code-level decisions discovered during BRD walkthrough.")

write_header_row(ws, 4,
                 ["#", "Decision", "Options", "Recommendation", "Decide By (Day)", "Decide By (Date)", "Status"],
                 widths=[5, 28, 32, 32, 14, 14, 14])

decisions = [
    (1, "AI provider for mapping",
     "OpenAI / Claude API / Local model",
     "Claude API — already used; cost-effective at this volume.", 1),
    (2, "Charting library",
     "Recharts / Chart.js / D3",
     "Recharts — declarative, works cleanly with React 19.", 1),
    (3, "Web scraping provider",
     "Apify (cloud) / Playwright (self) / Both",
     "Both — Apify for quick setups, Playwright for complex flows.", 7),
    (4, "Auth strategy",
     "Local (JWT) / SSO (Azure AD) / Both",
     "Local JWT first; pluggable adapter for SSO in Phase 3.", 1),
    (5, "Help content authoring",
     "Markdown in repo / CMS / Wiki",
     "Markdown in repo; index built at build time (lunr.js).", 28),
    (6, "Email transport",
     "SMTP / SES / SendGrid",
     "Start SMTP via env; swap for SES in Phase 4 if volume warrants.", 38),
    (7, "JS sandbox for transformations",
     "vm2 / QuickJS / WASM",
     "QuickJS — modern, sandboxed, low overhead, in worker thread.", 25),
    (8, "Hosting model",
     "Local server / Docker / Cloud (Azure/AWS)",
     "Start local + Docker for portability; cloud as Phase 4 outcome.", P3_LAST_DAY + 1),
    (9, "First non-Jira / non-SharePoint connector pair to demo",
     "PostgreSQL → TFS / Excel → Holiday Tracker / Dynamics → SharePoint",
     "PostgreSQL → TFS — both new, exercises Studio + new dest path.", 11),
    (10, "Multi-tenancy go-live",
     "Phase 2 (with auth) / Phase 3 (with client apps)",
     "Phase 3 — keep Phase 2 single-tenant to ship fast; org isolation lands with client apps.", P2_LAST_DAY),
    (11, "AI Help Assistant scope",
     "Library-only / + Screen state / + Action history",
     "Full context (library + state + history) at Phase 3 launch.", P2_LAST_DAY + 20),
    (12, "Mobile PWA depth",
     "Read-only / Read+limited write / Full parity",
     "Read-only at launch; expand based on actual mobile usage.", P3_LAST_DAY + 1),
]

for i, (num, dec, opts, rec, day) in enumerate(decisions):
    row = 5 + i
    is_alt = (i % 2 == 0)
    cal = P2_START + timedelta(days=day - 1)
    cells = [num, dec, opts, rec, day, cal, "Not Started"]
    for ci, val in enumerate(cells, start=1):
        c = ws.cell(row=row, column=ci, value=val)
        c.font = FONT_BODY
        c.alignment = ALIGN_LEFT
        c.border = border_all
        if is_alt:
            c.fill = FILL_ALT
    ws.cell(row=row, column=6).number_format = "yyyy-mm-dd"

last_d = 4 + len(decisions)
status_conditional(ws, "G", 5, last_d)
add_status_validation(ws, "G", 5, last_d)
# Decision status options: Not Started / In Progress / Done — same vocab as elsewhere

ws.freeze_panes = "A5"

# Save
wb.save(str(OUT))
print(f"WROTE: {OUT}")
print(f"SIZE:  {OUT.stat().st_size:,} bytes")
print(f"Phase 2 ends: Day {P2_LAST_DAY}")
print(f"Phase 3 ends: Day {P3_LAST_DAY}")
print(f"Phase 4 ends: Day {P4_LAST_DAY}")
print(f"Total span:   {P4_LAST_DAY} working days  ({P2_START} -> {P2_START + timedelta(days=P4_LAST_DAY-1)})")
