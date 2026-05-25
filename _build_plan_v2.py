"""
Synapse / XnHub Project Plan v2 — Excel Builder
Generates Synapse_Project_Plan_v2.xlsx with 6 tabs:
Cover, WBS, Milestones, Effort Summary, Sprint Plan, Risks & Open Items
"""

from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.formatting.rule import CellIsRule, FormulaRule, DataBarRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation

OUT = r"D:\Nalashaa_Work\AI_Work\synapse-fullstack\Synapse_Project_Plan_v2.1.xlsx"

# =============== BRAND ===============
NAVY = "1B2A5B"          # Nalashaa navy
LIGHT_NAVY = "E8ECF5"
ACCENT = "F26B21"        # orange accent
GREEN = "2E7D32"
RED = "C62828"
AMBER = "F9A825"
GRAY = "78909C"
DARK_GRAY = "455A64"
BLUE_STATUS = "1565C0"
LIGHT_GRAY = "F2F2F2"
WHITE = "FFFFFF"
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=WHITE)
TITLE_FONT = Font(name="Calibri", size=20, bold=True, color=NAVY)
SUBTITLE_FONT = Font(name="Calibri", size=12, bold=True, color=DARK_GRAY)
BODY_FONT = Font(name="Calibri", size=10)
BODY_BOLD = Font(name="Calibri", size=10, bold=True)
SMALL_FONT = Font(name="Calibri", size=9, color=DARK_GRAY)
NAVY_FILL = PatternFill("solid", fgColor=NAVY)
LIGHT_NAVY_FILL = PatternFill("solid", fgColor=LIGHT_NAVY)
ACCENT_FILL = PatternFill("solid", fgColor=ACCENT)
SUBTOTAL_FILL = PatternFill("solid", fgColor="D7DCEA")
THIN = Side(border_style="thin", color="B0BEC5")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

START_DATE = date(2026, 5, 25)
SPRINT_DAYS = 14

def sprint_start(s):
    return START_DATE + timedelta(days=SPRINT_DAYS * s)

def sprint_end(s):
    # 2 weeks - 1 day; assume Mon start, Fri 2nd week end -> +11 working days; here use cal end Sun (2 wks - 1)
    return sprint_start(s) + timedelta(days=SPRINT_DAYS - 3)  # Fri of week 2

# =============== WORKBOOK ===============
wb = Workbook()

# ===========================================================
# TAB 1 — COVER
# ===========================================================
ws = wb.active
ws.title = "Cover"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 32
ws.column_dimensions["C"].width = 90

# Title band
ws.merge_cells("B2:C2")
ws["B2"] = "Synapse / XnHub Integration Platform"
ws["B2"].font = TITLE_FONT
ws["B2"].alignment = LEFT

ws.merge_cells("B3:C3")
ws["B3"] = "Project Plan v2  ·  WBS · Milestones · Effort · Sprints · Risks"
ws["B3"].font = SUBTITLE_FONT

# Meta block
meta = [
    ("Generated",        date.today().isoformat()),
    ("Owner",            "Kiran Suvarna (Nalashaa)"),
    ("Client / Audience","Internal — Nalashaa Org-Level Integration Platform"),
    ("Code path",        r"D:\Nalashaa_Work\AI_Work\synapse-fullstack"),
    ("Source docs",      "Synapse Architecture v4 · Synapse LLD · Developer Handbook · XnHub Product Charter v2"),
    ("Plan version",     "v2.0"),
]
r = 5
for k, v in meta:
    ws.cell(r, 2, k).font = BODY_BOLD
    ws.cell(r, 3, v).font = BODY_FONT
    ws.cell(r, 2).alignment = LEFT
    ws.cell(r, 3).alignment = LEFT
    r += 1

# Assumptions
r += 1
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
ws.cell(r, 2, "Assumptions (redirect any of these and the plan recalculates)").font = SUBTITLE_FONT
ws.cell(r, 2).fill = LIGHT_NAVY_FILL
r += 1
assumptions = [
    ("Backend stack",        ".NET 8 (matches Architecture / LLD / Handbook verbatim). Python (FastAPI + Celery) is the documented alternative — swap the BE owner-role tasks if changed."),
    ("Scope",                "MVP (Phase 1) + Phase 2 (Intelligence & Monitoring) from XnHub Charter §8. Phase 3 / 4 tracked as roadmap only."),
    ("Vibecoding multiplier","0.5x of traditional hand-coded estimate. Vibe Hours column = Trad Hours × 0.5. Change cell B3 on the WBS tab to re-flow the entire plan."),
    ("Team",                 "1 BA · 1 FE Dev · 1 BE Dev · 1 QA · 0.25 DevOps. Capacity baseline: 160 hrs/person/month (8 hrs × 20 working days)."),
    ("Cadence",              "10 sprints of 2 weeks. Sprint 0 starts Mon 2026-05-25. Phase 2 UAT signoff target Fri 2026-11-13."),
    ("Frontend",             "React 19 + Vite at packages/frontend is KEPT AS-IS. Backend rewrite only. UI work = wire mocks → real API + new Login/Studio backend."),
]
for label, body in assumptions:
    ws.cell(r, 2, "▸ " + label).font = BODY_BOLD
    ws.cell(r, 3, body).font = BODY_FONT
    ws.cell(r, 2).alignment = LEFT_TOP
    ws.cell(r, 3).alignment = LEFT_TOP
    ws.row_dimensions[r].height = 32
    r += 1

# Codebase audit summary
r += 1
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
ws.cell(r, 2, "Existing codebase audit (PRODUCT_STATUS.md, 2026-04-22)").font = SUBTITLE_FONT
ws.cell(r, 2).fill = LIGHT_NAVY_FILL
r += 1
audit = [
    ("Frontend (React 19 + Vite)", "13 pages — all COMPLETE on UI. All use MOCK DATA except Wizard Step 2 (real API). Login screen NOT BUILT."),
    ("Backend (Node/Express)",     "FULLY BUILT: Jira (Red Gold REST + Flatiron Playwright/MFA), SharePoint Push (Graph API), 35-field mapper, AES-256-GCM vault, Scheduler (BullMQ), Sync delta, Run mgmt."),
    ("Backend gaps",               "NOT BUILT: Auth/JWT, Multi-tenancy enforcement, Alerts dispatch, Audit logging, Connector Studio backend, AI auto-map, Trading Console drill-down, Master Entity Catalog."),
    ("Migration approach",         "Port FULLY BUILT modules first (Sprints 1-3) → reach parity with Node version (M2) → then build NOT BUILT modules on top of new .NET hub."),
]
for label, body in audit:
    ws.cell(r, 2, "▸ " + label).font = BODY_BOLD
    ws.cell(r, 3, body).font = BODY_FONT
    ws.cell(r, 2).alignment = LEFT_TOP
    ws.cell(r, 3).alignment = LEFT_TOP
    ws.row_dimensions[r].height = 36
    r += 1

# Reading guide
r += 1
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
ws.cell(r, 2, "Tab guide").font = SUBTITLE_FONT
ws.cell(r, 2).fill = LIGHT_NAVY_FILL
r += 1
tabs = [
    ("WBS",            "Every task. Update Status, % Complete, and Delivered Date per task. On-Time and Days Slip auto-calculate."),
    ("Milestones",     "10 milestones (M0–M9) with target dates. Update Actual Date when achieved; Slip Days auto-calculates."),
    ("Effort Summary", "Live pivot of hours by Phase × Role. Reads directly from WBS — never edit cells here, edit WBS."),
    ("Sprint Plan",    "10 sprints with goals, demo items, capacity, and loaded utilization. Loaded Hours and Utilization auto-calculate."),
    ("Risks & Open",   "Charter §11 risks + .NET migration risks + 3 open decisions. Update Status as items close."),
]
for label, body in tabs:
    ws.cell(r, 2, label).font = BODY_BOLD
    ws.cell(r, 3, body).font = BODY_FONT
    ws.cell(r, 2).alignment = LEFT_TOP
    ws.cell(r, 3).alignment = LEFT_TOP
    r += 1

# Status legend
r += 1
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
ws.cell(r, 2, "Status legend (matches WBS conditional formatting)").font = SUBTITLE_FONT
ws.cell(r, 2).fill = LIGHT_NAVY_FILL
r += 1
legend = [
    ("Not Started", GRAY),
    ("In Progress", BLUE_STATUS),
    ("Blocked",     RED),
    ("Done",        GREEN),
    ("Cancelled",   DARK_GRAY),
]
for label, color in legend:
    ws.cell(r, 2, label).font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    ws.cell(r, 2).fill = PatternFill("solid", fgColor=color)
    ws.cell(r, 2).alignment = CENTER
    ws.cell(r, 2).border = BORDER
    r += 1

# ===========================================================
# TAB 2 — WBS
# ===========================================================
wbs = wb.create_sheet("WBS")
wbs.sheet_view.showGridLines = False

# Vibe multiplier control cell at top
wbs["A1"] = "Vibe multiplier:"
wbs["A1"].font = BODY_BOLD
wbs["A1"].alignment = Alignment(horizontal="right", vertical="center")
wbs["B1"] = 0.5
wbs["B1"].font = BODY_BOLD
wbs["B1"].fill = PatternFill("solid", fgColor="FFF2CC")
wbs["B1"].alignment = CENTER
wbs["B1"].number_format = "0.00"
wbs["B1"].border = BORDER
wbs["C1"] = "← change this cell to re-flow all Vibe Hours"
wbs["C1"].font = SMALL_FONT

# Header row at row 3
HEADERS = [
    "WBS ID","Phase","Workstream","Task","Owner Role",
    "Trad Hours","Vibe Hours","Sprint","Status","% Complete",
    "Original ETA","Current ETA","Delivered Date","On-Time?","Days Slip",
    "Notes / Dependencies"
]
for i, h in enumerate(HEADERS, 1):
    c = wbs.cell(3, i, h)
    c.font = HEADER_FONT
    c.fill = NAVY_FILL
    c.alignment = CENTER
    c.border = BORDER

# Column widths
widths = [10, 12, 22, 60, 12, 12, 12, 9, 13, 12, 13, 13, 14, 12, 11, 50]
for i, w in enumerate(widths, 1):
    wbs.column_dimensions[get_column_letter(i)].width = w
wbs.row_dimensions[3].height = 32
wbs.freeze_panes = "E4"

# ---- WBS DATA ----
# (Phase, Workstream, Task, Owner, Trad Hours, Sprint, ETA-offset-from-sprint-start in days, Notes)
# We'll set Original ETA = sprint_start(sprint) + offset days, with Fri-of-sprint as default
def eta_for(sprint, offset_days=10):
    """Default: Fri of week 2 of the sprint."""
    return sprint_start(sprint) + timedelta(days=offset_days)

WBS = []

# ---- PHASE 0 — Foundation & Setup (Sprint 0) ----
P0 = "Phase 0"
WBS += [
    (P0, "Setup",    "Stand up .NET 8 solution skeleton (Synapse.Core, Synapse.Host, Synapse.Tests) per LLD §2.1",                "BE",  16, 0, "Branch off main; reference LLD project dependency graph."),
    (P0, "Setup",    "Add Synapse.Connectors.Jira and Synapse.Connectors.SharePoint empty projects + ProjectReferences",          "BE",   4, 0, "Mirror Node module boundaries."),
    (P0, "Setup",    "Drizzle → EF Core schema migration (13 tables → C# entities, snake_case keep)",                             "BE",  24, 0, "Preserve org_id everywhere; keep app + jira_data schemas."),
    (P0, "DevOps",   "docker-compose: keep Postgres + Redis, replace Node service with .NET API service",                         "DevOps", 8, 0, "Two service blocks: api + playwright-worker."),
    (P0, "DevOps",   "GitHub Actions CI: dotnet build + test + Playwright install + frontend build",                              "DevOps", 12, 0, "Cache nuget + npm; matrix linux + windows."),
    (P0, "DevOps",   "Dockerfile for .NET API + Dockerfile.playwright for browser worker (separate image)",                       "DevOps", 8, 0, "Per CLAUDE.md §10 — Playwright workers need browser binaries."),
    (P0, "BA",       "Baseline requirements doc — map Charter §5 modules to MVP+Phase 2 backlog with acceptance criteria",        "BA",  16, 0, "Single source of truth for sprint planning."),
    (P0, "BA",       "Acceptance criteria template (Given/When/Then) + UAT script template",                                      "BA",   8, 0, "Reuse across all sprint demos."),
    (P0, "QA",       "Test strategy doc — unit (xUnit) / integration / E2E (Playwright) / perf / security split per Handbook §9", "QA",  12, 0, "Sets the test pyramid."),
    (P0, "Setup",    "Dev environment + onboarding doc (refresh existing CLAUDE.md for .NET path)",                               "BE",   8, 0, "Replace Node-specific steps."),
]

# ---- PHASE 1 — MVP (Sprints 1-5) ----
P1 = "Phase 1"

# Sprint 1 — Hub core abstractions + persistence
WBS += [
    (P1, "Hub Core",    "Implement ISourceConnector, IDestinationConnector, ITransformStep interfaces (LLD §6)",             "BE",  16, 1, "Three interfaces — the only contract surface."),
    (P1, "Hub Core",    "MessageEnvelope record + SHA-256 ComputeChecksum + IsChecksumValid (LLD §6, Handbook §4.1)",        "BE",  16, 1, "Immutable record; thread-safe Payload Clone()."),
    (P1, "Hub Core",    "IntegrationBus + bounded System.Threading.Channels intake + per-subscription channels",             "BE",  24, 1, "Backpressure naturally via Channel.Writer.WriteAsync."),
    (P1, "Hub Core",    "SubscriptionRegistry loader (subscriptions.json) + Subscription model",                             "BE",  16, 1, "Immutable post-load."),
    (P1, "Persistence", "EF Core HubDbContext + 4 entity types: InboxEntry, OutboxEntry, DeadLetterEntry, IdempotencyEntry", "BE",  20, 1, "Schema per LLD §8."),
    (P1, "Persistence", "InboxRepository + OutboxRepository + DeadLetterRepository + IdempotencyRepository",                 "BE",  24, 1, "Scoped lifetime; IDbContextFactory."),
    (P1, "QA",          "Unit tests: envelope checksum round-trip, channel backpressure, repository upsert",                 "QA",  16, 1, "Cover all 4 abstractions before connectors land."),
    (P1, "BA",          "Sprint 1 user-story refinement + acceptance criteria for hub core",                                 "BA",  10, 1, "Hub-engine stories: \"as a router, I can fan-out to N subs\"."),
]

# Sprint 2 — Router + workers + transform + DLQ + Polly
WBS += [
    (P1, "Hub Core",    "RouterService (BackgroundService): reads intake channel, fans out to subscription channels",        "BE",  20, 2, "Single loop; sole intake reader."),
    (P1, "Hub Core",    "SubscriptionWorkerOrchestrator + SubscriptionWorker per Subscription (Serial mode first)",          "BE",  24, 2, "Other modes deferred until Phase 2 perf."),
    (P1, "Hub Core",    "TransformPipeline executor (LLD §4.6) — chain ITransformStep by config order",                      "BE",  16, 2, "Unknown StepId: warn + skip."),
    (P1, "Reliability", "Polly ResiliencePipeline per destination: retry 3x exp+jitter + circuit breaker 0.5/30s + bulkhead","BE",  16, 2, "Per LLD §11.1 + Architecture §11."),
    (P1, "Reliability", "Dead Letter Queue auto-replay BackgroundService (5-min scan) + Poisoned state alerting",            "BE",  20, 2, "Per Architecture §11.2."),
    (P1, "Reliability", "Manual + bulk DLQ replay endpoints",                                                                "BE",  10, 2, "POST /hub/dlq/{id}/replay + bulk by destination."),
    (P1, "Hub Core",    "DI wiring: ServiceCollectionExtensions.AddSynapseCore() per LLD §7 table",                          "BE",  10, 2, "Lifetimes per LLD table 5."),
    (P1, "QA",          "Integration tests: end-to-end fake source → router → transform → fake destination → inbox=Done",    "QA",  20, 2, "Use Handbook §9.2 pipeline fakes pattern."),
    (P1, "BA",          "Sprint 2 demo prep + stakeholder walkthrough of hub engine",                                        "BA",   8, 2, "Demo M1 milestone candidate."),
]

# Sprint 3 — Port Jira connectors + SharePoint
WBS += [
    (P1, "Connector",   "Port RedGoldApiClient (REST + Basic Auth + p-limit) to .NET HttpClient + Polly",                    "BE",  24, 3, "Direct translation; preserve pagination loop."),
    (P1, "Connector",   "Port RedGoldDataExtractor JQL strategy + worklog expand → MessageEnvelope publish",                  "BE",  20, 3, "Status-changed JQL; worklog within report period."),
    (P1, "Connector",   "Port FlatironScraper to Microsoft.Playwright (.NET) + TOTP via Otp.NET",                            "BE",  32, 3, "MFA + storageState persistence (8hr); headed for debug."),
    (P1, "Connector",   "Port FlatironDataExtractor label-pattern fetch (patch_*, tempapp_*, nalashaa_*) + Testmo session",  "BE",  24, 3, "Per CLAUDE.md §5.2 label conventions."),
    (P1, "Connector",   "Port JiraTicketNormalizer (Flatiron + Red Gold → NormalizedJiraTicket)",                            "BE",  16, 3, "Pure mapping; covered by fixtures."),
    (P1, "Connector",   "Port SharePoint OAuth2 client-credentials auth + token cache + Graph site/list resolve",            "BE",  16, 3, "Microsoft.Graph SDK."),
    (P1, "Connector",   "Port SharePoint Push (create/update list items + 3-layer dedup + batch=20 + progress)",             "BE",  24, 3, "Preserve pushLog → jiraItemCache → SP query order."),
    (P1, "Connector",   "Port 35-field Jira→SharePoint mapper + derived fields (CycleTimeDays, IsOverdue, etc.)",            "BE",  20, 3, "Bring fixtures from Node tests."),
    (P1, "QA",          "Live integration test: Red Gold credential → real ticket pull → SharePoint test list write",        "QA",  16, 3, "Requires test Jira + SP tenant."),
    (P1, "QA",          "Live integration test: Flatiron MFA flow with headed Playwright + session restore",                 "QA",  16, 3, "Capture screenshots on failure."),
    (P1, "BA",          "Acceptance test for Jira → SharePoint parity (records-in = records-out, no dupes)",                 "BA",  10, 3, "Defines M2 milestone gate."),
]

# Sprint 4 — Vault + Scheduler + Run mgmt + API parity + RBAC schema
WBS += [
    (P1, "Vault",       "Port CredentialService (AES-256-GCM, per-org key) to .NET — KeyDerivation.HKDF",                    "BE",  16, 4, "Encrypted payload shape unchanged."),
    (P1, "Vault",       "Credentials API: POST /credentials, GET /credentials, GET /credentials/:id (metadata only)",        "BE",  12, 4, "Never return plaintext."),
    (P1, "Scheduler",   "Port SchedulerService from BullMQ → Quartz.NET cron with dynamic CRUD",                              "BE",  24, 4, "Persistent job store on Postgres."),
    (P1, "Runtime",     "Run management: RunRecord + status state machine (pending/running/success/error)",                  "BE",  16, 4, "Preserve records_in/records_out semantics."),
    (P1, "Runtime",     "Worker hosting: HostedService that owns Quartz + Playwright queues (concurrency=1 for browser)",    "BE",  20, 4, "Per CLAUDE.md §2.3."),
    (P1, "API",         "API parity: /integrations, /jira, /sharepoint, /push, /sync, /connected — Express routes → ASP.NET",  "BE",  32, 4, "Zod → FluentValidation. Same response envelope."),
    (P1, "Security",    "RBAC schema (roles enum) + users/orgs seed + JWT bearer middleware (org_id claim)",                 "BE",  20, 4, "Cookie + Bearer; refresh tokens."),
    (P1, "Security",    "Row-level multi-tenancy: query filters by org_id on every entity (global EF filter)",               "BE",  16, 4, "HasQueryFilter pattern."),
    (P1, "Alerts",      "Basic alerts: failure detection in BackgroundService → write Alert row (in-app only)",              "BE",  16, 4, "Email/Slack deferred to Phase 2."),
    (P1, "Audit",       "Audit log middleware: write to audit_log on auth-sensitive routes (CRUD vault/integrations/users)", "BE",  12, 4, "Append-only; before/after diff."),
    (P1, "QA",          "Unit tests: AES-256-GCM round-trip + JWT claims + RLS query filter",                                 "QA",  16, 4, "Wrong-key throws, expired token rejects."),
    (P1, "QA",          "API contract tests vs old Node API (parity oracle)",                                                "QA",  20, 4, "Postman/RestAssured replay."),
    (P1, "BA",          "RBAC permission matrix sign-off (Admin / Designer / Operator / Viewer per Charter §5.8.2 table)",   "BA",  10, 4, "Permission grid is the source of truth."),
]

# Sprint 5 — Frontend wire-up + Login + Studio + MVP UAT
WBS += [
    (P1, "Frontend",    "Login screen + Auth.js integration (calls .NET JWT endpoint, stores in HttpOnly cookie)",           "FE",  16, 5, "First time real auth ships."),
    (P1, "Frontend",    "Wire Dashboard: remove mockDashboardTiles, fetch /integrations + /runs, real KPI cards",            "FE",  12, 5, "Add loading + empty states."),
    (P1, "Frontend",    "Wire Registry: real /integrations list + sparklines from /runs",                                    "FE",  10, 5, "Replace mock integrations.js."),
    (P1, "Frontend",    "Wire Monitor: real /runs + expandable row with /runs/:id detail (payload + mapping trace)",         "FE",  14, 5, "Largest table — paginate."),
    (P1, "Frontend",    "Wire Alerts: real /alerts list + acknowledge action POST /alerts/:id/ack",                          "FE",   8, 5, "Severity sort."),
    (P1, "Frontend",    "Wire Vault: real /credentials (masked) + add/rotate flow with eye-icon reveal (10s auto-hide)",     "FE",  12, 5, "Per Charter §5.9."),
    (P1, "Frontend",    "Wire Connected: real /connected list + schedule modal hooked to PATCH schedule",                    "FE",  10, 5, "Cron presets unchanged."),
    (P1, "Frontend",    "Wire Push: real POST /push call + dedup modal from server response",                                "FE",  10, 5, "Surface server-side dedup."),
    (P1, "Frontend",    "Wire Admin: real /users CRUD + role edit + activate/deactivate",                                    "FE",  12, 5, "Soft delete only."),
    (P1, "Frontend",    "Wizard Step 4 (Mapping): replace redirect stub with inline canvas + commit to integration",         "FE",  16, 5, "Largest wizard fix."),
    (P1, "Frontend",    "Wizard Step 6 (Test & Deploy): actual test execution via /integrations/:id/run + publish flow",     "FE",  12, 5, "Hard gate per Charter §5.2."),
    (P1, "Frontend",    "Studio page: enable Test Connection + Publish buttons + wire to /connectors endpoints",             "FE",  16, 5, "Companion backend in BE row below."),
    (P1, "Frontend",    "Charts: replace hardcoded SVG with Recharts wired to /runs aggregates (volume + error trend)",      "FE",  14, 5, "Use SIS-P1 5-second rule layout."),
    (P1, "Frontend",    "Loading states + error toasts + Zod form validation across all pages",                              "FE",  16, 5, "Cross-cutting cleanup."),
    (P1, "Backend",     "Connectors CRUD API (/connectors POST/GET/PUT/publish) for Studio page",                            "BE",  20, 5, "Use existing connectors table."),
    (P1, "QA",          "E2E test suite (Playwright Test): login → create integration → run → view in monitor",              "QA",  24, 5, "Per CLAUDE.md §12."),
    (P1, "QA",          "Regression: full Jira→SP run, schedule fire, vault rotate, dedup behaviour",                        "QA",  16, 5, "Smoke + critical paths."),
    (P1, "QA",          "Security review: OWASP top 10 + credential vault threat model + JWT settings",                      "QA",  16, 5, "Hard gate before UAT."),
    (P1, "BA",          "MVP UAT planning + test scenarios + stakeholder schedule",                                          "BA",  16, 5, "Coordinate with all module owners."),
    (P1, "BA",          "MVP UAT execution + signoff coordination",                                                          "BA",  20, 5, "M5 milestone owner."),
    (P1, "DevOps",      "Stage env deploy: containers + Postgres + Redis + smoke health check",                              "DevOps",12, 5, "Stage = production rehearsal."),
    (P1, "DevOps",      "Structured logging (Serilog) + correlation IDs + log shipping",                                     "DevOps", 8, 5, "Per Architecture §6 envelope CorrelationId."),
]

# ---- PHASE 2 — Intelligence & Monitoring (Sprints 6-10) ----
P2 = "Phase 2"

# Sprint 6 — AI Auto-Map + NL Transform
WBS += [
    (P2, "AI",          "AI Auto-Mapping service: schema-aware prompt to Claude API + confidence score per pair",            "BE",  32, 6, "Per Charter §5.3.2 + D2 decision."),
    (P2, "AI",          "Persist suggestion review state (suggested / confirmed / rejected / adjusted) per integration",     "BE",  12, 6, "New table or JSONB on integration row."),
    (P2, "AI",          "NL Transformation builder: NL → expression (Roslyn script for .NET; sandboxed)",                    "BE",  32, 6, "Sandboxed compilation; allow-list types."),
    (P2, "Frontend",    "Mapping Canvas: AI Auto-Map button + dashed-line suggestions + confidence badges + confirm/reject", "FE",  24, 6, "Per Charter §5.3.2 visual spec."),
    (P2, "Frontend",    "Transformation slide-up panel: presets + NL input + AI-generated script preview",                   "FE",  20, 6, "Per Charter §5.3.3."),
    (P2, "QA",          "AI test set: 50-pair mapping fixtures, measure auto-map accuracy vs >70% target",                   "QA",  16, 6, "Charter §10 success metric."),
    (P2, "BA",          "AI mapping acceptance criteria + fallback rules (always require human confirmation)",               "BA",  10, 6, "Charter §11 risk mitigation."),
]

# Sprint 7 — Trading Console + Health Dashboard
WBS += [
    (P2, "Monitoring",  "Trading Network Console backend: /messages query with payload + transform trace per message",       "BE",  24, 7, "Heavy read — paginate aggressively."),
    (P2, "Frontend",    "Trading Console UI: table + expandable rows (JSON syntax-highlight, mapping trace, error stack)",   "FE",  24, 7, "Per Charter §5.6."),
    (P2, "Monitoring",  "Health Dashboard backend: /metrics aggregations (per-adapter status, volume, error trend)",         "BE",  20, 7, "Materialized view if needed."),
    (P2, "Frontend",    "Health Dashboard UI: adapter tiles + volume stacked-bar + error trend line + drill-down",           "FE",  24, 7, "Real data; replace mock dashboardTiles.js."),
    (P2, "Frontend",    "Real-time refresh (SSE) for Trading Console + Health Dashboard pages",                              "FE",  16, 7, "Pause button per Charter §5.6."),
    (P2, "QA",          "Perf test: 4-worker Parallel mode @ 100 msg/s for 1 hour — verify 3x throughput vs Serial",         "QA",  16, 7, "Per Architecture §7.1 claim."),
    (P2, "QA",          "Trading Console: payload search + filter combo regression",                                         "QA",  10, 7, "Common ops use case."),
    (P2, "BA",          "Demo: Trading Console + Health Dashboard to stakeholders (M7 milestone)",                            "BA",   8, 7, "Sprint 7 demo."),
]

# Sprint 7 ADD — SharePoint Source side of SP→DB pipeline
WBS += [
    (P2, "SP→DB",       "SharePointListSourceConnector — Graph /items/delta polling + per-(site,list) deltaLink token store", "BE",  32, 7, "Reuses app.sync_state table; topic sharepoint.list.{id}.changed."),
    (P2, "SP→DB",       "SharePointFieldIntrospector — read SP list column metadata for Mapping Canvas left pane",            "BE",  12, 7, "Returns (name, internalName, sharePointType, indexed, required)."),
    (P2, "SP→DB",       "Webhook intake placeholder: POST /webhooks/sharepoint stub (validate handshake, log, 200)",          "BE",   8, 7, "Architecture-ready, feature-flagged off. Phase 2 wire-up later."),
    (P2, "SP→DB",       "DbConnectionCredential type in Vault + AES-256-GCM payload shape + POST /credentials/:id/test",       "BE",  12, 7, "host/port/db/user/pwd/schema/sslMode/engine."),
    (P2, "Frontend",    "Studio: add SharePoint List (source) + PostgreSQL + SQL Server connector tiles + icons",              "FE",   8, 7, "Plug into existing Studio category grid."),
    (P2, "Frontend",    "Wizard Step 2: DB connection credential form (host/port/db/user/pwd/schema/ssl) + Test Connection",   "FE",  12, 7, "Mirrors Jira/SP credential forms."),
    (P2, "QA",          "Unit tests: deltaLink persistence, resume after restart, deletion-tombstone handling",                "QA",  10, 7, "Critical — token loss = full resync."),
    (P2, "BA",          "SP-source + DB-dest acceptance criteria + connector spec sign-off",                                   "BA",  10, 7, "Defines M8 gate for this slice."),
]

# Sprint 8 — Master Entity Catalog + Connectors (incl. DB-dest side of SP→DB) + Scraping
WBS += [
    (P2, "Catalog",     "Master Entity Catalog backend: entity registry + cross-adapter usage + field-level stats",          "BE",  28, 8, "Per Charter §5.5."),
    (P2, "Frontend",    "Master Entity Catalog UI: tree nav + field table + usage bars + cross-reference links",             "FE",  24, 8, "Replace catalog page hardcoded data."),
    (P2, "SP→DB",       "DbDestinationConnectorBase — abstract: batching, Polly wrap, transaction-per-batch, IDestinationConnector contract", "BE",  20, 8, "Strategy pattern; engine-specific writers hang off this."),
    (P2, "SP→DB",       "PostgresWriter — Npgsql + INSERT ... ON CONFLICT (sp_item_id) DO UPDATE upsert",                    "BE",  20, 8, "Idempotent on SP item ID natural key."),
    (P2, "SP→DB",       "SqlServerWriter — Microsoft.Data.SqlClient + MERGE INTO ... USING ... ON ... upsert",                "BE",  20, 8, "MSSQL MERGE semantics; preserve UTC."),
    (P2, "SP→DB",       "DbSchemaIntrospector (PG via information_schema + MSSQL via sys.columns) — design-time call",       "BE",  16, 8, "Returns columns, types, nullable, default."),
    (P2, "SP→DB",       "SharePointToSqlTypeMapper — per-engine type lookup (Person→2 cols, Lookup→2 cols, Choice-multi→jsonb/json)", "BE",  24, 8, "~30 fixture rows, one of the highest test-density modules."),
    (P2, "SP→DB",       "DdlDiffService — diff(spFields × mapping × dbColumns) → CREATE/ALTER statements + explain text",    "BE",  20, 8, "Never executes; only proposes."),
    (P2, "SP→DB",       "DdlApplyService — execute approved DDL in transaction, audit_log entry, prod-env 2nd-person guard", "BE",  12, 8, "Hard gate before Publish."),
    (P2, "Frontend",    "Mapping Canvas: render destination columns from DbSchemaIntrospector + flag unmapped/missing cols",  "FE",  16, 8, "Replaces destination panel mock data."),
    (P2, "Frontend",    "Wizard Step 6: DDL Preview panel — render statements + Apply DDL confirm modal + block Publish until clear", "FE",  16, 8, "Preview-before-commit per Charter UX principle."),
    (P2, "Connector",   "File Share connector (CSV/Excel source/dest with delimiter + sheet picker)",                        "BE",  24, 8, "Per Charter Phase 2 scope."),
    (P2, "Scraping",    "Apify cloud actor connector (config + run-trigger + result fetch)",                                 "BE",  20, 8, "Per Charter §5.1.1 web scraping (a)."),
    (P2, "Scraping",    "Playwright Automation connector (recorded/scripted flow runner + JSON result mapper)",              "BE",  24, 8, "Per Charter §5.1.1 web scraping (b)."),
    (P2, "Scraping",    "Scraping connector breakage alert (screenshot-diff on Playwright failure)",                         "BE",  10, 8, "Charter §5.11.1."),
    (P2, "QA",          "E2E test: SP list → MessageEnvelope → Postgres table — happy path, dedup on replay, type coercion", "QA",  16, 8, "Critical-path test."),
    (P2, "QA",          "E2E test: SP list → MessageEnvelope → SQL Server table — happy path, dedup on replay, type coercion","QA",  14, 8, "Mirror of PG suite."),
    (P2, "QA",          "Type-mapper fixtures (~30 rows: Person, Lookup, Choice-multi, Managed Metadata, DateTime, Yes/No)", "QA",  16, 8, "Run against both PG + MSSQL."),
    (P2, "QA",          "DDL diff acceptance: missing-col, type-mismatch, narrower-than-source — verify proposal + apply",   "QA",  12, 8, "Block-Publish behaviour test."),
    (P2, "QA",          "Webhook stub: validation handshake response + payload log, no bus side-effect (Phase 2 wire later)","QA",   4, 8, "Architecture-ready confirmation."),
    (P2, "QA",          "Connector regression: File Share + Apify + Playwright — source → transform → dest happy + error",   "QA",  18, 8, "3 connectors × 2 paths."),
    (P2, "BA",          "DDL approval UX walkthrough + sign-off on operator confirm flow",                                   "BA",   8, 8, "Sensitive — DBAs will scrutinise."),
    (P2, "BA",          "Entity merge UX spec + impact-analysis acceptance criteria",                                        "BA",  12, 8, "Defer merge impl to Phase 3."),
]

# Sprint 9 — Full Alerting + Task escalation + Cron UI + RBAC overrides
WBS += [
    (P2, "Alerts",      "Email alert dispatcher (SendGrid) + per-user notification prefs (immediate/digest/off)",            "BE",  20, 9, "Per Charter §5.11.2."),
    (P2, "Alerts",      "Escalation chains: timer-based escalate-to-admin + auto-pause on 2nd unacknowledged",               "BE",  16, 9, "Per Charter §5.11.3."),
    (P2, "Alerts",      "Threshold-based custom alerts engine (per-adapter rule store + evaluator)",                         "BE",  20, 9, "Volume drop, error rate, etc."),
    (P2, "Frontend",    "Alerts UI: notification bell + feed + detail panel + acknowledge/escalate actions",                 "FE",  20, 9, "Replace mock alerts.js."),
    (P2, "Workflow",    "Task escalation system: Operator→Designer task with auto-attached context (mapping/error)",         "BE",  16, 9, "Per Charter §5.6 + §5.3.3."),
    (P2, "Frontend",    "Cron builder UI in Wizard Step 5 (visual cron with next-run preview)",                              "FE",  16, 9, "Per Charter §5.2 step 5."),
    (P2, "Security",    "Resource-level RBAC overrides + admin UI to grant per-adapter permission exceptions",               "BE",  20, 9, "Per Charter §5.8.3."),
    (P2, "Frontend",    "Audit log viewer UI in Admin (filterable by user, action, entity, date)",                           "FE",  14, 9, "Append-only view."),
    (P2, "Frontend",    "Contextual help: first-time walkthroughs + inline form hints + tooltip library",                    "FE",  16, 9, "Per Charter §5.10.2."),
    (P2, "QA",          "Alerts E2E: failure → in-app + email → escalation timer → auto-pause",                              "QA",  16, 9, "Time-traveled tests."),
    (P2, "QA",          "Phase 2 regression + security re-review (new APIs, AI sandbox boundaries)",                         "QA",  20, 9, "Pre-UAT gate."),
    (P2, "BA",          "Phase 2 UAT planning + execution + signoff",                                                        "BA",  24, 9, "M9 milestone owner."),
    (P2, "DevOps",      "Production deploy + monitoring (Serilog → Seq or ELK + uptime probe)",                              "DevOps", 16, 9, "Go-live readiness."),
]

# ---- Write rows ----
ROW_START = 4
row = ROW_START
phase_first_row = {}
phase_last_row = {}
for i, (phase, ws_name, task, owner, trad, sprint, notes) in enumerate(WBS, 1):
    wid = f"W-{i:03d}"
    if phase not in phase_first_row:
        phase_first_row[phase] = row
    phase_last_row[phase] = row

    eta = eta_for(sprint)
    wbs.cell(row, 1, wid).font = BODY_FONT
    wbs.cell(row, 2, phase).font = BODY_FONT
    wbs.cell(row, 3, ws_name).font = BODY_FONT
    wbs.cell(row, 4, task).font = BODY_FONT
    wbs.cell(row, 5, owner).font = BODY_FONT
    wbs.cell(row, 6, trad).font = BODY_FONT
    # Vibe Hours formula
    wbs.cell(row, 7, f"=F{row}*$B$1").font = BODY_FONT
    wbs.cell(row, 8, sprint).font = BODY_FONT
    wbs.cell(row, 9, "Not Started").font = BODY_FONT
    wbs.cell(row, 10, 0).font = BODY_FONT
    wbs.cell(row, 11, eta).font = BODY_FONT
    wbs.cell(row, 12, eta).font = BODY_FONT  # Current ETA default = original
    wbs.cell(row, 13, None).font = BODY_FONT  # Delivered Date
    # On-Time? formula
    wbs.cell(row, 14, f'=IF(M{row}="","",IF(M{row}<=K{row},"On-Time","Delayed"))').font = BODY_FONT
    # Days Slip formula
    wbs.cell(row, 15, f'=IF(M{row}="","",NETWORKDAYS(K{row},M{row})-1)').font = BODY_FONT
    wbs.cell(row, 16, notes).font = BODY_FONT

    # Formatting
    for col in range(1, 17):
        c = wbs.cell(row, col)
        c.border = BORDER
        c.alignment = LEFT_TOP if col in (3, 4, 16) else CENTER
    wbs.cell(row, 10).number_format = "0%"
    wbs.cell(row, 10).value = 0
    wbs.cell(row, 11).number_format = "yyyy-mm-dd"
    wbs.cell(row, 12).number_format = "yyyy-mm-dd"
    wbs.cell(row, 13).number_format = "yyyy-mm-dd"
    wbs.cell(row, 7).number_format = "0.0"
    wbs.row_dimensions[row].height = 28
    row += 1

last_data_row = row - 1

# Subtotal rows per phase
for phase in ("Phase 0", "Phase 1", "Phase 2"):
    fr = phase_first_row[phase]
    lr = phase_last_row[phase]
    wbs.cell(row, 2, f"{phase} subtotal").font = BODY_BOLD
    wbs.cell(row, 6, f"=SUM(F{fr}:F{lr})").font = BODY_BOLD
    wbs.cell(row, 7, f"=SUM(G{fr}:G{lr})").font = BODY_BOLD
    wbs.cell(row, 6).number_format = "#,##0"
    wbs.cell(row, 7).number_format = "#,##0.0"
    for col in range(1, 17):
        wbs.cell(row, col).fill = SUBTOTAL_FILL
        wbs.cell(row, col).border = BORDER
    row += 1

grand_row = row
wbs.cell(row, 2, "GRAND TOTAL").font = HEADER_FONT
wbs.cell(row, 6, f"=SUM(F{ROW_START}:F{last_data_row})").font = HEADER_FONT
wbs.cell(row, 7, f"=SUM(G{ROW_START}:G{last_data_row})").font = HEADER_FONT
wbs.cell(row, 6).number_format = "#,##0"
wbs.cell(row, 7).number_format = "#,##0.0"
for col in range(1, 17):
    wbs.cell(row, col).fill = NAVY_FILL
    wbs.cell(row, col).font = HEADER_FONT
    wbs.cell(row, col).border = BORDER

# Data validations
dv_status = DataValidation(type="list", formula1='"Not Started,In Progress,Blocked,Done,Cancelled"', allow_blank=True)
dv_status.add(f"I{ROW_START}:I{last_data_row}")
wbs.add_data_validation(dv_status)

dv_pct = DataValidation(type="list", formula1='"0,0.25,0.5,0.75,1"', allow_blank=True)
dv_pct.add(f"J{ROW_START}:J{last_data_row}")
wbs.add_data_validation(dv_pct)

dv_role = DataValidation(type="list", formula1='"BA,FE,BE,QA,DevOps,Backend,Frontend"', allow_blank=True)
dv_role.add(f"E{ROW_START}:E{last_data_row}")
wbs.add_data_validation(dv_role)

# Conditional formatting for Status column
status_colors = {
    "Not Started": GRAY,
    "In Progress": BLUE_STATUS,
    "Blocked":     RED,
    "Done":        GREEN,
    "Cancelled":   DARK_GRAY,
}
for status, hex_color in status_colors.items():
    fill = PatternFill("solid", fgColor=hex_color)
    font = Font(color=WHITE, bold=True, size=10)
    rule = CellIsRule(operator="equal", formula=[f'"{status}"'], fill=fill, font=font)
    wbs.conditional_formatting.add(f"I{ROW_START}:I{last_data_row}", rule)

# Conditional formatting for On-Time?
wbs.conditional_formatting.add(
    f"N{ROW_START}:N{last_data_row}",
    CellIsRule(operator="equal", formula=['"On-Time"'], fill=PatternFill("solid", fgColor=GREEN), font=Font(color=WHITE, bold=True, size=10)),
)
wbs.conditional_formatting.add(
    f"N{ROW_START}:N{last_data_row}",
    CellIsRule(operator="equal", formula=['"Delayed"'], fill=PatternFill("solid", fgColor=RED), font=Font(color=WHITE, bold=True, size=10)),
)

# Data bar on % Complete
wbs.conditional_formatting.add(
    f"J{ROW_START}:J{last_data_row}",
    DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color="4472C4", showValue=True),
)

# ===========================================================
# TAB 3 — MILESTONES
# ===========================================================
ms = wb.create_sheet("Milestones")
ms.sheet_view.showGridLines = False

ms_headers = ["Milestone ID","Milestone","Phase","Target Date","Status","Owner","Acceptance Criteria","Dependencies","Actual Date","Slip Days"]
for i, h in enumerate(ms_headers, 1):
    c = ms.cell(1, i, h)
    c.font = HEADER_FONT
    c.fill = NAVY_FILL
    c.alignment = CENTER
    c.border = BORDER

widths_ms = [12, 55, 10, 14, 14, 14, 60, 30, 14, 11]
for i, w in enumerate(widths_ms, 1):
    ms.column_dimensions[get_column_letter(i)].width = w
ms.row_dimensions[1].height = 30
ms.freeze_panes = "A2"

MILESTONES = [
    ("M0","Sprint 0 complete — solution scaffolded, CI green, EF migrations applied",                            "P0", date(2026,6,5),  "Kiran",
        "dotnet build green; dotnet test 0 failures; docker compose up healthy; CI pipeline runs in <8 min.",
        "—"),
    ("M1","Hub engine working — first message Inbox → Channel → Router → Outbox → DLQ on forced failure",        "P1", date(2026,6,19), "BE Dev",
        "Fake source emits envelope; checksum valid; PENDING row in inbox; routes to subscription; DLQ entry on forced 500.",
        "M0"),
    ("M2","Jira → SharePoint parity with Node version (Red Gold + Flatiron + 35-field mapper)",                  "P1", date(2026,7,17), "BE Dev",
        "Live Red Gold + Flatiron pull → SP write; records-in = records-out; zero duplicates across re-run; matches Node-version output byte-for-byte for 100 sample tickets.",
        "M1"),
    ("M3","Frontend wired end-to-end — all 13 pages on real API, mocks removed",                                 "P1", date(2026,7,31), "FE Dev",
        "Mock data imports removed from Dashboard/Registry/Monitor/Alerts/Vault/Connected/Push/Admin/Studio; Wizard Step 4 & 6 functional; Recharts wired.",
        "M2"),
    ("M4","Auth + Multi-tenancy + RBAC live (4 roles + JWT + row-level org filter)",                              "P1", date(2026,8,14), "BE Dev",
        "Login flow works; cross-org data isolation verified by penetration test; permission matrix per Charter §5.8.2 enforced; audit log writes on sensitive actions.",
        "M3"),
    ("M5","MVP UAT signoff (Phase 1 complete)",                                                                  "P1", date(2026,8,28), "BA",
        "Stakeholder UAT scripts passed; Charter §10 metrics met (Operator setup <30 min, 0% hardcoded creds, test-before-publish 100%); production deploy approved.",
        "M4"),
    ("M6","AI Auto-Map + NL Transformation GA",                                                                   "P2", date(2026,9,25), "BE Dev",
        "Auto-map >70% accuracy on 50-pair fixture (Charter §10); NL → expression sandbox compiles & runs; human-confirmation gate present.",
        "M5"),
    ("M7","Trading Console + Health Dashboard GA (real data, real-time)",                                        "P2", date(2026,10,9), "FE Dev",
        "Per-message trace expandable; one-click Create Task; Charts/SVG removed; 4-worker Parallel mode demonstrates 3x throughput vs Serial.",
        "M6"),
    ("M8","Master Entity Catalog + SP→DB (PG+MSSQL) + File/Scraping connectors GA",                              "P2", date(2026,10,30),"BE Dev",
        "Entity tree + field-level usage live; SharePoint list (delta-poll) → Postgres + SQL Server with DDL-diff approve flow; CSV + Apify + Playwright connectors pass happy + error paths; webhook intake stub returns 200 + logs; scraping breakage screenshot-diff alert wired.",
        "M7"),
    ("M9","Phase 2 UAT signoff + production deploy",                                                             "P2", date(2026,11,13),"BA",
        "Full Phase 2 acceptance script passed; alerting end-to-end (email + escalation + auto-pause); RBAC resource overrides verified; on-call runbook handed over.",
        "M8"),
]

for i, (mid, m, phase, target, owner, ac, dep) in enumerate(MILESTONES, 2):
    ms.cell(i, 1, mid).font = BODY_BOLD
    ms.cell(i, 2, m).font = BODY_FONT
    ms.cell(i, 3, phase).font = BODY_FONT
    ms.cell(i, 4, target).font = BODY_FONT
    ms.cell(i, 4).number_format = "yyyy-mm-dd"
    ms.cell(i, 5, "Not Started").font = BODY_FONT
    ms.cell(i, 6, owner).font = BODY_FONT
    ms.cell(i, 7, ac).font = BODY_FONT
    ms.cell(i, 8, dep).font = BODY_FONT
    ms.cell(i, 9, None).number_format = "yyyy-mm-dd"
    ms.cell(i, 10, f'=IF(I{i}="","",NETWORKDAYS(D{i},I{i})-1)').font = BODY_FONT
    for col in range(1, 11):
        ms.cell(i, col).border = BORDER
        ms.cell(i, col).alignment = LEFT_TOP if col in (2, 7, 8) else CENTER
    ms.row_dimensions[i].height = 60

dv_ms_status = DataValidation(type="list", formula1='"Not Started,In Progress,At Risk,Achieved,Missed"', allow_blank=True)
dv_ms_status.add(f"E2:E{len(MILESTONES)+1}")
ms.add_data_validation(dv_ms_status)

for status, hex_color in {
    "Not Started": GRAY, "In Progress": BLUE_STATUS, "At Risk": AMBER, "Achieved": GREEN, "Missed": RED
}.items():
    ms.conditional_formatting.add(
        f"E2:E{len(MILESTONES)+1}",
        CellIsRule(operator="equal", formula=[f'"{status}"'], fill=PatternFill("solid", fgColor=hex_color), font=Font(color=WHITE, bold=True, size=10)),
    )

# ===========================================================
# TAB 4 — EFFORT SUMMARY
# ===========================================================
es = wb.create_sheet("Effort Summary")
es.sheet_view.showGridLines = False

es_headers = ["Phase","BA Hours","FE Dev Hours","BE Dev Hours","QA Hours","DevOps Hours","Total Vibe Hours","Total Person-Days","% of Total"]
for i, h in enumerate(es_headers, 1):
    c = es.cell(2, i, h)
    c.font = HEADER_FONT
    c.fill = NAVY_FILL
    c.alignment = CENTER
    c.border = BORDER

widths_es = [16, 14, 16, 16, 14, 16, 18, 18, 12]
for i, w in enumerate(widths_es, 1):
    es.column_dimensions[get_column_letter(i)].width = w
es.row_dimensions[2].height = 30

es["A1"] = "Effort by Phase × Role (live from WBS — never edit cells here)"
es["A1"].font = SUBTITLE_FONT
es.merge_cells("A1:I1")

# Helper to build SUMIFS against WBS tab
def sumifs_formula(phase, role):
    # G = Vibe Hours, B = Phase, E = Owner Role
    return f'=SUMIFS(WBS!G{ROW_START}:G{last_data_row},WBS!B{ROW_START}:B{last_data_row},"{phase}",WBS!E{ROW_START}:E{last_data_row},"{role}")'

phases = [("Phase 0","Phase 0"),("Phase 1","Phase 1"),("Phase 2","Phase 2")]
es_row = 3
phase_rows = []
for label, key in phases:
    es.cell(es_row, 1, label).font = BODY_BOLD
    es.cell(es_row, 2, sumifs_formula(key, "BA"))
    es.cell(es_row, 3, sumifs_formula(key, "FE"))
    # BE column combines BE + Backend keyword (for any rows tagged "Backend")
    es.cell(es_row, 4, f'={sumifs_formula(key,"BE")[1:]}+{sumifs_formula(key,"Backend")[1:]}')
    es.cell(es_row, 5, sumifs_formula(key, "QA"))
    es.cell(es_row, 6, sumifs_formula(key, "DevOps"))
    es.cell(es_row, 7, f"=SUM(B{es_row}:F{es_row})")
    es.cell(es_row, 8, f"=G{es_row}/8")
    es.cell(es_row, 9, f"=IFERROR(G{es_row}/$G$6,0)")
    for col in range(2, 10):
        es.cell(es_row, col).number_format = "#,##0.0"
        if col == 9:
            es.cell(es_row, col).number_format = "0.0%"
    for col in range(1, 10):
        es.cell(es_row, col).border = BORDER
        es.cell(es_row, col).alignment = CENTER
    phase_rows.append(es_row)
    es_row += 1

# Grand total row
es.cell(es_row, 1, "GRAND TOTAL").font = HEADER_FONT
for col in range(2, 8):
    letter = get_column_letter(col)
    es.cell(es_row, col, f"=SUM({letter}3:{letter}{es_row-1})")
    es.cell(es_row, col).number_format = "#,##0.0"
es.cell(es_row, 8, f"=G{es_row}/8")
es.cell(es_row, 8).number_format = "#,##0.0"
es.cell(es_row, 9, "100.0%")
for col in range(1, 10):
    es.cell(es_row, col).fill = NAVY_FILL
    es.cell(es_row, col).font = HEADER_FONT
    es.cell(es_row, col).border = BORDER
    es.cell(es_row, col).alignment = CENTER

# Capacity check block
cap_start = es_row + 3
es.cell(cap_start, 1, "Capacity check — over-allocation flag").font = SUBTITLE_FONT
es.merge_cells(start_row=cap_start, start_column=1, end_row=cap_start, end_column=9)
es.cell(cap_start, 1).fill = LIGHT_NAVY_FILL

cap_hdr = ["Role","Total Vibe Hours","Person-Days","Months (÷160)","Calendar Months Avail","Capacity Flag"]
for i, h in enumerate(cap_hdr, 1):
    c = es.cell(cap_start + 1, i, h)
    c.font = HEADER_FONT
    c.fill = NAVY_FILL
    c.alignment = CENTER
    c.border = BORDER

# Project runs roughly Sprint 0 (2026-05-25) to end of Sprint 9 + UAT (~2026-11-13) → 5.6 months
project_months = 5.6
roles_cap = [
    ("BA",    f"=SUM(B3:B{es_row-1})"),
    ("FE",    f"=SUM(C3:C{es_row-1})"),
    ("BE",    f"=SUM(D3:D{es_row-1})"),
    ("QA",    f"=SUM(E3:E{es_row-1})"),
    ("DevOps",f"=SUM(F3:F{es_row-1})"),
]
for i, (role, formula) in enumerate(roles_cap, cap_start + 2):
    es.cell(i, 1, role).font = BODY_BOLD
    es.cell(i, 2, formula).number_format = "#,##0.0"
    es.cell(i, 3, f"=B{i}/8").number_format = "#,##0.0"
    es.cell(i, 4, f"=B{i}/160").number_format = "0.00"
    es.cell(i, 5, project_months).number_format = "0.0"
    es.cell(i, 6, f'=IF(D{i}>E{i},"OVER-ALLOCATED",IF(D{i}>E{i}*0.85,"AT-LIMIT","OK"))').font = BODY_BOLD
    for col in range(1, 7):
        es.cell(i, col).border = BORDER
        es.cell(i, col).alignment = CENTER

last_cap_row = cap_start + 1 + len(roles_cap)
# Conditional formatting on capacity flag
es.conditional_formatting.add(
    f"F{cap_start+2}:F{last_cap_row}",
    CellIsRule(operator="equal", formula=['"OVER-ALLOCATED"'], fill=PatternFill("solid", fgColor=RED), font=Font(color=WHITE, bold=True, size=10)),
)
es.conditional_formatting.add(
    f"F{cap_start+2}:F{last_cap_row}",
    CellIsRule(operator="equal", formula=['"AT-LIMIT"'], fill=PatternFill("solid", fgColor=AMBER), font=Font(color=WHITE, bold=True, size=10)),
)
es.conditional_formatting.add(
    f"F{cap_start+2}:F{last_cap_row}",
    CellIsRule(operator="equal", formula=['"OK"'], fill=PatternFill("solid", fgColor=GREEN), font=Font(color=WHITE, bold=True, size=10)),
)

# Footnote
es.cell(last_cap_row + 2, 1, "Note: BE Dev Hours sums both 'BE' and 'Backend' role tags on WBS. Adjust headcount if any role is OVER-ALLOCATED.").font = SMALL_FONT
es.merge_cells(start_row=last_cap_row + 2, start_column=1, end_row=last_cap_row + 2, end_column=9)

# ===========================================================
# TAB 5 — SPRINT PLAN
# ===========================================================
sp = wb.create_sheet("Sprint Plan")
sp.sheet_view.showGridLines = False

sp_headers = ["Sprint","Start Date","End Date","Sprint Goal","Demo Items","Capacity Hours","Loaded Hours","Utilization %","Milestone Hit"]
for i, h in enumerate(sp_headers, 1):
    c = sp.cell(1, i, h)
    c.font = HEADER_FONT
    c.fill = NAVY_FILL
    c.alignment = CENTER
    c.border = BORDER

widths_sp = [10, 14, 14, 50, 50, 16, 16, 14, 16]
for i, w in enumerate(widths_sp, 1):
    sp.column_dimensions[get_column_letter(i)].width = w
sp.row_dimensions[1].height = 30
sp.freeze_panes = "A2"

# 4 people × 8 hrs × 10 working days × 2 wks = 4 × 80 = 320 hrs (we use 4 people across BA/FE/BE/QA, DevOps adds 0.25)
sprint_capacity = 4 * 80  # 320 hrs/sprint baseline

SPRINTS = [
    (0, "Foundation",                "Sprint 0 — Foundation & Setup",
        ".NET solution skeleton · EF Core schema · docker-compose · CI green · BA baseline doc · QA test strategy", "M0"),
    (1, "Hub Core I",                "Sprint 1 — Hub Core Abstractions + Persistence",
        "ISourceConnector/IDest/ITransform interfaces · MessageEnvelope · IntegrationBus + Channels · 4 repositories", "—"),
    (2, "Hub Core II + Reliability", "Sprint 2 — Router, Workers, Transform, Polly, DLQ",
        "Router fan-out · SubscriptionWorker · TransformPipeline · Polly retry + breaker · DLQ auto-replay", "M1"),
    (3, "Connector Port",            "Sprint 3 — Port Jira (Red Gold + Flatiron) + SharePoint + Mapper",
        "Live Jira→SP via .NET; parity to Node baseline on 100-ticket fixture", "M2"),
    (4, "Auth + API Parity",         "Sprint 4 — Vault, Scheduler, Run mgmt, API parity, RBAC schema",
        "Quartz cron · CredentialService AES-256 · /integrations + /jira + /sp routes · JWT + RLS · audit middleware", "—"),
    (5, "FE Wire-Up + MVP UAT",      "Sprint 5 — Frontend wire-up + Studio backend + MVP UAT",
        "13 pages wired (no mocks) · Login · Wizard 4+6 · Studio · Recharts · UAT signoff", "M3, M4, M5"),
    (6, "AI Auto-Map",               "Sprint 6 — AI Auto-Mapping + NL Transformation",
        "Claude-API mapper with >70% accuracy · NL→Roslyn sandbox · Canvas auto-map UX", "M6"),
    (7, "Trading Console + SP Source","Sprint 7 — Trading Console + Health Dashboard + SP source side of SP→DB",
        "Per-message trace · adapter tiles + volume + error trend · 4-worker perf demo · SharePointListSourceConnector (delta) + DbConnectionCredential vault type · webhook stub", "M7"),
    (8, "Catalog + SP→DB + Scraping","Sprint 8 — Master Entity Catalog + DB destination (PG+MSSQL) + DDL diff + File/Scraping",
        "Entity tree · PostgresWriter + SqlServerWriter (UPSERT) · DdlDiffService + Apply flow · SP→PG + SP→MSSQL E2E · CSV + Apify + Playwright connectors", "M8"),
    (9, "Alerts + Phase 2 UAT",      "Sprint 9 — Full Alerting + Task Escalation + RBAC overrides + Phase 2 UAT",
        "Email alerts + escalation + auto-pause · cron UI · resource overrides · audit viewer · prod deploy", "M9"),
]

for i, (s, _, goal, demo, ms_hit) in enumerate(SPRINTS, 2):
    sp.cell(i, 1, f"Sprint {s}").font = BODY_BOLD
    sp.cell(i, 2, sprint_start(s)).number_format = "yyyy-mm-dd"
    sp.cell(i, 3, sprint_end(s)).number_format = "yyyy-mm-dd"
    sp.cell(i, 4, goal).font = BODY_FONT
    sp.cell(i, 5, demo).font = BODY_FONT
    sp.cell(i, 6, sprint_capacity).font = BODY_BOLD
    sp.cell(i, 7, f'=SUMIFS(WBS!G{ROW_START}:G{last_data_row},WBS!H{ROW_START}:H{last_data_row},{s})').font = BODY_FONT
    sp.cell(i, 7).number_format = "#,##0.0"
    sp.cell(i, 8, f"=IFERROR(G{i}/F{i},0)").font = BODY_FONT
    sp.cell(i, 8).number_format = "0.0%"
    sp.cell(i, 9, ms_hit).font = BODY_FONT
    for col in range(1, 10):
        sp.cell(i, col).border = BORDER
        sp.cell(i, col).alignment = LEFT_TOP if col in (4, 5) else CENTER
    sp.row_dimensions[i].height = 56

# Conditional formatting on utilization
sp.conditional_formatting.add(
    f"H2:H{len(SPRINTS)+1}",
    CellIsRule(operator="greaterThan", formula=["1"], fill=PatternFill("solid", fgColor=RED), font=Font(color=WHITE, bold=True, size=10)),
)
sp.conditional_formatting.add(
    f"H2:H{len(SPRINTS)+1}",
    CellIsRule(operator="between", formula=["0.85","1"], fill=PatternFill("solid", fgColor=AMBER), font=Font(color=WHITE, bold=True, size=10)),
)
sp.conditional_formatting.add(
    f"H2:H{len(SPRINTS)+1}",
    CellIsRule(operator="lessThan", formula=["0.85"], fill=PatternFill("solid", fgColor=GREEN), font=Font(color=WHITE, bold=True, size=10)),
)

sp.cell(len(SPRINTS)+3, 1, "Capacity assumption: 4 people × 8 hrs × 10 working days = 320 hrs/sprint. Adjust col F if team size changes.").font = SMALL_FONT
sp.merge_cells(start_row=len(SPRINTS)+3, start_column=1, end_row=len(SPRINTS)+3, end_column=9)

# ===========================================================
# TAB 6 — RISKS & OPEN ITEMS
# ===========================================================
rk = wb.create_sheet("Risks & Open Items")
rk.sheet_view.showGridLines = False

rk_headers = ["ID","Type","Description","Probability","Impact","Severity","Mitigation / Action","Owner","Due Date","Status"]
for i, h in enumerate(rk_headers, 1):
    c = rk.cell(1, i, h)
    c.font = HEADER_FONT
    c.fill = NAVY_FILL
    c.alignment = CENTER
    c.border = BORDER

widths_rk = [8, 10, 60, 12, 10, 11, 60, 14, 14, 14]
for i, w in enumerate(widths_rk, 1):
    rk.column_dimensions[get_column_letter(i)].width = w
rk.row_dimensions[1].height = 30
rk.freeze_panes = "A2"

RISKS = [
    # Charter §11 risks
    ("R-01","Risk","External API access not obtainable for some target systems (auth, rate limits, IP allow-list)",
        "M","H","Three-tier fallback: browser-passthrough auth → Apify cloud → Playwright self-hosted. Flag scraping connectors with bot icon in registry.", "BA", date(2026,6,15), "Open"),
    ("R-02","Risk","Web scraping connectors break when target UI changes",
        "H","M","Screenshot-diff alert on scrape failure; Designer notified immediately; flag scraping connectors as high-maintenance.", "BE Dev", date(2026,9,1), "Open"),
    ("R-03","Risk","Scope creep — Charter Phase 3/4 items pulled into MVP",
        "H","H","Lock scope per phase; change-request process; weekly scope review against Charter §8 phase boundaries.", "BA", date(2026,5,29), "Open"),
    ("R-04","Risk","AI field mapper produces poor suggestions in domain-specific schemas",
        "M","M","Always require human confirmation; log confirm/reject decisions to tune prompt; graceful fallback to manual map.", "BE Dev", date(2026,9,18), "Open"),
    ("R-05","Risk","Non-technical Operators find Mapping Canvas too complex",
        "M","H","Two rounds of moderated UX testing in Sprint 5 + 8; progressive disclosure; one-click Request Designer Help.", "BA", date(2026,8,15), "Open"),
    ("R-06","Risk","Credential vault security vulnerability",
        "L","C","AES-256-GCM + security review (M5 gate); never log raw values; OWASP top-10 covered in Sprint 5 QA task.", "BE Dev", date(2026,8,28), "Open"),
    # .NET migration specific
    ("R-07","Risk","Node→.NET port effort underestimated (Playwright/Flatiron MFA esp.)",
        "M","H","Validate with thin Sprint 1 slice; rebaseline B1 vibe multiplier and Trad Hours if Sprint 3 spillover >20%.", "Kiran", date(2026,7,17), "Open"),
    ("R-08","Risk","Microsoft.Playwright (.NET) reliability for long-running MFA sessions",
        "M","M","Run 24-hour headless soak in Sprint 3; keep storageState persistence + re-login fallback identical to Node version.", "BE Dev", date(2026,7,17), "Open"),
    ("R-09","Risk","0.5x Vibecoding multiplier may not hold for complex hub logic (channels, Polly, RLS)",
        "M","H","Track actual vs. Vibe Hours per task; review at end of Sprint 2; rebaseline B1 cell if variance >25%.", "Kiran", date(2026,6,19), "Open"),
    ("R-10","Risk","Frontend mock-to-real swap surfaces hidden contract gaps (shape mismatch, missing fields)",
        "M","M","Generate TypeScript types from .NET DTOs (NSwag) before FE wire-up; contract tests in Sprint 4 (W-060).", "FE Dev", date(2026,7,24), "Open"),
    ("R-11","Risk","OAuth2/MFA flows for SharePoint + Flatiron sensitive to client config drift",
        "M","H","Per-client smoke job after deploy; alert on auth-error spike; document tenant prerequisites in onboarding doc.", "BE Dev", date(2026,8,14), "Open"),
    # Decisions
    ("D-01","Decision","Final backend choice — .NET 8 vs Python (FastAPI)",
        "—","H","Default .NET (matches all design docs). Re-affirm at Sprint 0 close; if changed, swap BE tasks to Python equivalents (FastAPI + SQLAlchemy + Celery + Playwright-Python).", "Kiran", date(2026,6,5), "Open"),
    ("D-02","Decision","AI provider for auto-mapping — Claude API / OpenAI / local model",
        "—","M","Bench-test top 2 candidates on a 50-pair fixture in Sprint 5; pick on cost × accuracy × latency.", "BE Dev", date(2026,9,4), "Open"),
    ("D-03","Decision","Hosting model — local / Docker / cloud (Azure or AWS)",
        "—","M","Start local + Docker; choose cloud target by Sprint 2 to inform DevOps work in Sprint 4-5.", "DevOps", date(2026,6,19), "Open"),
    ("D-04","Decision","First two production clients/connectors after MVP",
        "—","M","Confirm with stakeholders by Sprint 1 demo. Drives Phase 2 connector priority (DB/File/MQ vs scraping).", "BA", date(2026,6,19), "Open"),
    ("D-05","Decision","Transformation scripting language — Roslyn C# / Python / JavaScript / DSL",
        "—","M","Default Roslyn C# (matches .NET host). Revisit if Python backend chosen (use RestrictedPython).", "BE Dev", date(2026,8,28), "Open"),
]

for i, (rid, rtype, desc, prob, impact, mit, owner, due, status) in enumerate(RISKS, 2):
    rk.cell(i, 1, rid).font = BODY_BOLD
    rk.cell(i, 2, rtype).font = BODY_FONT
    rk.cell(i, 3, desc).font = BODY_FONT
    rk.cell(i, 4, prob).font = BODY_FONT
    rk.cell(i, 5, impact).font = BODY_FONT
    # Severity = Probability score × Impact score; simple grid via lookup
    rk.cell(i, 6, f'=IFERROR(IF(OR(D{i}="—",E{i}="—"),"—",CHOOSE(MATCH(D{i},{{"L","M","H"}},0),"L","M","H")&"·"&CHOOSE(MATCH(E{i},{{"L","M","H","C"}},0),"L","M","H","C")),"—")').font = BODY_FONT
    rk.cell(i, 7, mit).font = BODY_FONT
    rk.cell(i, 8, owner).font = BODY_FONT
    rk.cell(i, 9, due).number_format = "yyyy-mm-dd"
    rk.cell(i, 10, status).font = BODY_FONT
    for col in range(1, 11):
        rk.cell(i, col).border = BORDER
        rk.cell(i, col).alignment = LEFT_TOP if col in (3, 7) else CENTER
    rk.row_dimensions[i].height = 48

# Data validation
dv_type = DataValidation(type="list", formula1='"Risk,Issue,Decision,Assumption"', allow_blank=True)
dv_type.add(f"B2:B{len(RISKS)+1}")
rk.add_data_validation(dv_type)

dv_prob = DataValidation(type="list", formula1='"L,M,H,—"', allow_blank=True)
dv_prob.add(f"D2:D{len(RISKS)+1}")
rk.add_data_validation(dv_prob)

dv_imp = DataValidation(type="list", formula1='"L,M,H,C,—"', allow_blank=True)
dv_imp.add(f"E2:E{len(RISKS)+1}")
rk.add_data_validation(dv_imp)

dv_rk_status = DataValidation(type="list", formula1='"Open,Mitigating,Monitoring,Closed,Escalated"', allow_blank=True)
dv_rk_status.add(f"J2:J{len(RISKS)+1}")
rk.add_data_validation(dv_rk_status)

# Status colors
for status, hex_color in {
    "Open": AMBER, "Mitigating": BLUE_STATUS, "Monitoring": "9C27B0", "Closed": GREEN, "Escalated": RED
}.items():
    rk.conditional_formatting.add(
        f"J2:J{len(RISKS)+1}",
        CellIsRule(operator="equal", formula=[f'"{status}"'], fill=PatternFill("solid", fgColor=hex_color), font=Font(color=WHITE, bold=True, size=10)),
    )

# Impact color
rk.conditional_formatting.add(
    f"E2:E{len(RISKS)+1}",
    CellIsRule(operator="equal", formula=['"C"'], fill=PatternFill("solid", fgColor=RED), font=Font(color=WHITE, bold=True, size=10)),
)
rk.conditional_formatting.add(
    f"E2:E{len(RISKS)+1}",
    CellIsRule(operator="equal", formula=['"H"'], fill=PatternFill("solid", fgColor=AMBER), font=Font(color=WHITE, bold=True, size=10)),
)

# Save
wb.save(OUT)
print(f"OK :: {OUT}")
print(f"WBS rows: {len(WBS)}, Milestones: {len(MILESTONES)}, Sprints: {len(SPRINTS)}, Risks/Decisions: {len(RISKS)}")
